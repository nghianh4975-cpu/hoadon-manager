# -*- coding: utf-8 -*-
"""
HOA DON WEB APP - Flask Application
"""

import os
import sqlite3
import uuid
import datetime
import base64
import json
from io import BytesIO
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify

try:
    from werkzeug.security import generate_password_hash, check_password_hash
except:
    import hashlib
    def generate_password_hash(pwd):
        return hashlib.sha256(pwd.encode()).hexdigest()
    def check_password_hash(hashed, pwd):
        return hashed == hashlib.sha256(pwd.encode()).hexdigest()

try:
    import cv2
    import pytesseract
    HAS_OCR = True
except:
    HAS_OCR = False

# ============================================================
# APP CONFIG
# ============================================================
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', str(uuid.uuid4()))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Use PostgreSQL on Render, SQLite when local
USE_PG = bool(os.environ.get('DATABASE_URL'))

if USE_PG:
    DATABASE = None  # PostgreSQL doesn't need local file
else:
    DATABASE = os.path.join(BASE_DIR, 'hoadon.db')

UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# Jinja2 filter
@app.template_filter('from_json')
def from_json(value):
    try:
        return json.loads(value) if value else []
    except:
        return []

# ============================================================
# DATABASE - PostgreSQL tren Render, SQLite khi local
# ============================================================

class HybridCursor:
    def __init__(self, cursor, use_pg):
        self._cursor = cursor
        self._use_pg = use_pg
        self._col_names = None
    def execute(self, sql, params=None):
        if params is not None:
            if self._use_pg:
                sql = sql.replace('?', '%s')
            self._cursor.execute(sql, params)
        else:
            self._cursor.execute(sql)
        if self._use_pg:
            self._col_names = [desc[0] for desc in self._cursor.description] if self._cursor.description else []
        return self
    def fetchone(self):
        row = self._cursor.fetchone()
        if row is None:
            return None
        if self._use_pg and self._col_names:
            d = dict(zip(self._col_names, row))
            class Row:
                def __init__(self, d):
                    self._d = d
                def __getitem__(self, k):
                    if isinstance(k, int):
                        return list(self._d.values())[k]
                    return self._d[k]
                def get(self, k, default=None):
                    return self._d.get(k, default)
                def keys(self):
                    return self._d.keys()
                def values(self):
                    return self._d.values()
                def items(self):
                    return self._d.items()
                def __len__(self):
                    return len(self._d)
                def __iter__(self):
                    return iter(self._d)
            return Row(d)
        return row
    def fetchall(self):
        rows = self._cursor.fetchall()
        if self._use_pg and self._col_names:
            def make_row(r):
                d = dict(zip(self._col_names, r))
                class Row:
                    def __init__(self, d):
                        self._d = d
                    def __getitem__(self, k):
                        if isinstance(k, int):
                            return list(self._d.values())[k]
                        return self._d[k]
                    def get(self, k, default=None):
                        return self._d.get(k, default)
                    def keys(self):
                        return self._d.keys()
                    def __len__(self):
                        return len(self._d)
                    def __iter__(self):
                        return iter(self._d)
                return Row(d)
            return [make_row(r) for r in rows]
        return rows
    def commit(self):
        self._cursor.connection.commit()
    def close(self):
        self._cursor.close()
    def __enter__(self):
        return self
    def __exit__(self, *args):
        self.close()
    @property
    def lastrowid(self):
        return getattr(self._cursor, 'lastrowid', None)

def get_db():
    if USE_PG:
        import psycopg2
        conn = psycopg2.connect(os.environ['DATABASE_URL'])
        conn.autocommit = True  # DDL can fail without aborting transaction
    else:
        conn = sqlite3.connect(DATABASE)
        conn.row_factory = sqlite3.Row

    class Wrapper:
        def __init__(self, conn, use_pg):
            self._conn = conn
            self._use_pg = use_pg
        def cursor(self):
            return HybridCursor(self._conn.cursor(), self._use_pg)
        def commit(self):
            self._conn.commit()
        def close(self):
            self._conn.close()
        def __enter__(self):
            return self
        def __exit__(self, *args):
            self.close()

    return Wrapper(conn, USE_PG)

def init_db():
    conn = get_db()
    c = conn.cursor()
    if USE_PG:
        # PostgreSQL schema
        c.execute('''CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username VARCHAR(255) UNIQUE NOT NULL,
            password VARCHAR(255) NOT NULL,
            role VARCHAR(50) NOT NULL DEFAULT 'ketoan',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS invoices (
            id SERIAL PRIMARY KEY,
            invoice_number VARCHAR(255),
            date VARCHAR(20),
            store_name VARCHAR(255),
            items TEXT,
            subtotal REAL DEFAULT 0,
            discount_percent REAL DEFAULT 0,
            tax_percent REAL DEFAULT 0,
            total REAL DEFAULT 0,
            notes TEXT,
            image_data TEXT,
            created_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS finances (
            id SERIAL PRIMARY KEY,
            type VARCHAR(20) NOT NULL,
            date VARCHAR(20) NOT NULL,
            amount REAL NOT NULL,
            category VARCHAR(255),
            reason TEXT,
            description TEXT,
            created_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS notifications (
            id SERIAL PRIMARY KEY,
            user_id INTEGER,
            message TEXT,
            invoice_id INTEGER,
            is_read INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS material_groups (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            color VARCHAR(20) DEFAULT '',
            sort_order INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS suppliers (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            phone VARCHAR(50),
            address TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS materials (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            unit VARCHAR(50) DEFAULT '',
            group_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS imports (
            id SERIAL PRIMARY KEY,
            date VARCHAR(20) NOT NULL,
            material_id INTEGER,
            supplier_id INTEGER,
            quantity REAL DEFAULT 0,
            unit_price REAL DEFAULT 0,
            total_price REAL DEFAULT 0,
            notes TEXT,
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            paid_amount REAL DEFAULT 0
        )''')
        try:
            c.execute('ALTER TABLE imports ADD COLUMN paid_amount REAL DEFAULT 0')
        except Exception:
            pass
        try:
            c.execute("ALTER TABLE imports ADD COLUMN unit TEXT DEFAULT ''")
        except Exception:
            pass
        try:
            c.execute('ALTER TABLE imports ADD COLUMN payment_count INTEGER DEFAULT 0')
        except Exception:
            pass
        try:
            c.execute("ALTER TABLE imports ADD COLUMN paid_details TEXT DEFAULT ''")
        except Exception:
            pass
        c.execute('''CREATE TABLE IF NOT EXISTS inventory (
            id SERIAL PRIMARY KEY,
            month VARCHAR(7) NOT NULL,
            material_id INTEGER,
            opening_stock REAL DEFAULT 0,
            import_qty REAL DEFAULT 0,
            closing_stock REAL DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('SELECT COUNT(*) FROM users')
        if c.fetchone()[0] == 0:
            c.execute('INSERT INTO users (username, password, role) VALUES (%s, %s, %s)',
                      ('admin', generate_password_hash('admin123'), 'admin'))
    else:
        # SQLite schema
        c.execute('''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'ketoan',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_number TEXT,
            date TEXT,
            store_name TEXT,
            items TEXT,
            subtotal REAL DEFAULT 0,
            discount_percent REAL DEFAULT 0,
            tax_percent REAL DEFAULT 0,
            total REAL DEFAULT 0,
            notes TEXT,
            image_data TEXT,
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS finances (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL,
            date TEXT NOT NULL,
            amount REAL NOT NULL,
            category TEXT,
            reason TEXT,
            description TEXT,
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            message TEXT,
            invoice_id INTEGER,
            is_read INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        # ---- Nhom hang truoc (vi materials tham chieu no) ----
        c.execute('''CREATE TABLE IF NOT EXISTS material_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            color TEXT DEFAULT '',
            sort_order INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT,
            address TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            unit TEXT DEFAULT '',
            group_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS imports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            material_id INTEGER,
            supplier_id INTEGER,
            quantity REAL DEFAULT 0,
            unit TEXT DEFAULT '',
            unit_price REAL DEFAULT 0,
            total_price REAL DEFAULT 0,
            notes TEXT,
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            paid_amount REAL DEFAULT 0
        )''')
        try:
            c.execute('ALTER TABLE imports ADD COLUMN paid_amount REAL DEFAULT 0')
        except Exception:
            pass
        try:
            c.execute("ALTER TABLE imports ADD COLUMN unit TEXT DEFAULT ''")
        except Exception:
            pass
        try:
            c.execute('ALTER TABLE imports ADD COLUMN payment_count INTEGER DEFAULT 0')
        except Exception:
            pass
        try:
            c.execute("ALTER TABLE imports ADD COLUMN paid_details TEXT DEFAULT ''")
        except Exception:
            pass
        c.execute('''CREATE TABLE IF NOT EXISTS inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            month TEXT NOT NULL,
            material_id INTEGER,
            opening_stock REAL DEFAULT 0,
            import_qty REAL DEFAULT 0,
            closing_stock REAL DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(month, material_id)
        )''')
        c.execute('SELECT COUNT(*) FROM users')
        if c.fetchone()[0] == 0:
            c.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                      ('admin', generate_password_hash('admin123'), 'admin'))

    conn.commit()
    conn.close()

# ============================================================
# AUTH
# ============================================================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'admin':
            flash('Ban khong co quyen!', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

# Context processor - tu dong truyen notif_count vao moi template
@app.context_processor
def inject_user():
    if 'user_id' not in session:
        return {'notif_count': 0}
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM notifications WHERE user_id=? AND is_read=0', (session['user_id'],))
    count = c.fetchone()[0]
    conn.close()
    return {'notif_count': count}

@app.before_request
def before():
    pass

# ============================================================
# ROUTES - AUTH
# ============================================================
@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT * FROM users WHERE username = ?', (username,))
        user = c.fetchone()
        conn.close()

        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            flash(f'Chao muon {user["username"]}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Dang nhap that bai!', 'danger')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current = request.form.get('current_password', '')
        new_pass = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')

        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT password FROM users WHERE id = ?', (session['user_id'],))
        user = c.fetchone()

        if not check_password_hash(user['password'], current):
            flash('Mat khau cu khong dung!', 'danger')
            conn.close()
            return render_template('change_password.html')

        if new_pass != confirm:
            flash('Mat khau moi khong khop!', 'danger')
            conn.close()
            return render_template('change_password.html')

        c.execute('UPDATE users SET password = ? WHERE id = ?',
                  (generate_password_hash(new_pass), session['user_id']))
        conn.commit()
        conn.close()
        flash('Doi mat khau thanh cong!', 'success')
        return redirect(url_for('dashboard'))

    return render_template('change_password.html')

# ============================================================
# ROUTES - DASHBOARD
# ============================================================
@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    c = conn.cursor()

    today = datetime.date.today().strftime('%Y-%m-%d')
    month_start = datetime.date.today().replace(day=1).strftime('%Y-%m-%d')

    c.execute('SELECT COUNT(*), COALESCE(SUM(total),0) FROM invoices WHERE date = ?', (today,))
    t = c.fetchone()
    today_count, today_total = t[0], t[1]

    c.execute('SELECT COUNT(*), COALESCE(SUM(total),0) FROM invoices WHERE date >= ?', (month_start,))
    t = c.fetchone()
    month_count, month_total = t[0], t[1]

    c.execute("SELECT COALESCE(SUM(amount),0) FROM finances WHERE type='revenue' AND date >= ?", (month_start,))
    month_rev = c.fetchone()[0]

    c.execute("SELECT COALESCE(SUM(amount),0) FROM finances WHERE type='expense' AND date >= ?", (month_start,))
    month_exp = c.fetchone()[0]

    c.execute('SELECT COUNT(*) FROM notifications WHERE user_id=? AND is_read=0', (session['user_id'],))
    notif_count = c.fetchone()[0]

    c.execute('SELECT * FROM invoices ORDER BY created_at DESC LIMIT 5')
    recent = c.fetchall()
    conn.close()

    return render_template('dashboard.html',
        today_count=today_count, today_total=today_total,
        month_count=month_count, month_total=month_total,
        month_rev=month_rev, month_exp=month_exp,
        profit=month_rev - month_exp,
        recent_invoices=recent)

# ============================================================
# ROUTES - INVOICES
# ============================================================
@app.route('/invoices')
@login_required
def invoices():
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    conn = get_db()
    c = conn.cursor()

    query = 'SELECT * FROM invoices WHERE 1=1'
    params = []
    if date_from:
        query += ' AND date >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND date <= ?'
        params.append(date_to)
    query += ' ORDER BY date DESC, id DESC'

    c.execute(query, params)
    invoices_list = c.fetchall()
    conn.close()

    return render_template('invoices.html', invoices=invoices_list,
                         date_from=date_from, date_to=date_to)

@app.route('/invoice/new', methods=['GET', 'POST'])
@login_required
def new_invoice():
    if request.method == 'POST':
        items_json = request.form.get('items_data', '[]')
        subtotal = float(request.form.get('subtotal', 0) or 0)
        discount = float(request.form.get('discount_percent', 0) or 0)
        tax = float(request.form.get('tax_percent', 0) or 0)
        total = subtotal * (1 - discount/100) * (1 + tax/100)

        # Image
        image_data = ''
        if 'invoice_image' in request.files:
            file = request.files['invoice_image']
            if file and file.filename:
                img_bytes = file.read()
                image_data = base64.b64encode(img_bytes).decode('utf-8')

        conn = get_db()
        c = conn.cursor()

        # Get invoice number
        c.execute('SELECT COUNT(*) FROM invoices')
        count = c.fetchone()[0]
        invoice_number = request.form.get('invoice_number') or f"HD{datetime.date.today().strftime('%Y%m%d')}{count+1:04d}"

        # Insert and get ID
        if USE_PG:
            c.execute('''INSERT INTO invoices
                (invoice_number, date, store_name, items, subtotal, discount_percent,
                 tax_percent, total, notes, image_data, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id''',
                (invoice_number,
                 request.form.get('invoice_date'),
                 request.form.get('store_name'),
                 items_json, subtotal, discount, tax, total,
                 request.form.get('notes', ''),
                 image_data,
                 session['username']))
            invoice_id = c.fetchone()[0]
        else:
            c.execute('''INSERT INTO invoices
                (invoice_number, date, store_name, items, subtotal, discount_percent,
                 tax_percent, total, notes, image_data, created_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (invoice_number,
                 request.form.get('invoice_date'),
                 request.form.get('store_name'),
                 items_json, subtotal, discount, tax, total,
                 request.form.get('notes', ''),
                 image_data,
                 session['username']))
            invoice_id = c.lastrowid

        conn.commit()
        conn.close()

        flash('Tao hoa don thanh cong!', 'success')
        return redirect(url_for('view_invoice', invoice_id=invoice_id))

    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM invoices')
    count = c.fetchone()[0]
    conn.close()
    invoice_number = f"HD{datetime.date.today().strftime('%Y%m%d')}{count+1:04d}"

    return render_template('invoice_form.html',
        invoice={'invoice_number': invoice_number,
                 'date': datetime.date.today().strftime('%Y-%m-%d')})

# Nhap text xuat hang loat
@app.route('/invoice/import-text', methods=['GET', 'POST'])
@login_required
def import_text():
    if request.method == 'POST':
        text = request.form.get('text_data', '')
        invoice_date = request.form.get('invoice_date', datetime.date.today().strftime('%Y-%m-%d'))
        store_name = request.form.get('store_name', '')
        num_cols = int(request.form.get('num_cols', 4) or 4)

        items = []
        grand_total = 0

        lines = text.strip().split('\n')
        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Skip total line
            lower_line = line.lower()
            if 'tong' in lower_line or 'thanh tien' in lower_line or 'total' in lower_line:
                continue

            # Split by |
            parts = None
            if '\t' in line:
                parts = [p.strip() for p in line.split('\t') if p.strip()]
            elif '|' in line:
                parts = [p.strip() for p in line.split('|') if p.strip()]
            else:
                # Try to find last number in line
                tokens = line.split()
                try:
                    price = float(tokens[-1].replace(',', '').replace('.', ''))
                    name = ' '.join(tokens[:-1]).strip()
                    parts = [name, '1', str(price), str(price)]
                except:
                    pass

            if parts and len(parts) >= 2:
                name = parts[0].strip()
                if not name:
                    continue

                if num_cols == 2:
                    # Ten | Gia (so luong = 1)
                    qty = 1
                    price = float(parts[1].replace(',', ''))
                else:
                    # Ten | SoLuong | DonGia [ | ThanhTien]
                    try:
                        qty = float(parts[1].replace(',', ''))
                    except:
                        qty = 1
                    try:
                        price = float(parts[2].replace(',', ''))
                    except:
                        price = 0

                total = qty * price
                items.append({
                    'name': name,
                    'quantity': qty,
                    'price': price,
                    'total': total
                })
                grand_total += total

        conn = get_db()
        c = conn.cursor()

        # Generate invoice number
        c.execute('SELECT COUNT(*) FROM invoices')
        count = c.fetchone()[0]
        invoice_number = f"HD{datetime.date.today().strftime('%Y%m%d')}{count+1:04d}"

        # Insert and get ID
        if USE_PG:
            c.execute('''INSERT INTO invoices
                (invoice_number, date, store_name, items, subtotal, total, notes, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id''',
                (invoice_number, invoice_date, store_name,
                 json.dumps(items), grand_total, grand_total,
                 f'Tu dong tao tu nhap text - {len(items)} san pham', session['username']))
            invoice_id = c.fetchone()[0]
        else:
            c.execute('''INSERT INTO invoices
                (invoice_number, date, store_name, items, subtotal, total, notes, created_by)
                VALUES (?,?,?,?,?,?,?,?)''',
                (invoice_number, invoice_date, store_name,
                 json.dumps(items), grand_total, grand_total,
                 f'Tu dong tao tu nhap text - {len(items)} san pham', session['username']))
            invoice_id = c.lastrowid

        conn.commit()
        conn.close()

        flash(f'Tao {len(items)} san pham thanh cong!', 'success')
        return redirect(url_for('view_invoice', invoice_id=invoice_id))

    return render_template('import_text.html',
        today_str=datetime.date.today().strftime('%Y-%m-%d'))

# Quet anh OCR
@app.route('/invoice/scan-image', methods=['GET', 'POST'])
@login_required
def scan_image():
    if request.method == 'POST':
        if 'invoice_image' not in request.files:
            flash('Vui long upload anh!', 'danger')
            return redirect(request.url)

        file = request.files['invoice_image']
        if not file.filename:
            flash('Vui long upload anh!', 'danger')
            return redirect(request.url)

        img_bytes = file.read()
        image_data = base64.b64encode(img_bytes).decode('utf-8')

        # OCR
        extracted_text = ''
        if HAS_OCR:
            try:
                nparr = BytesIO(img_bytes)
                img_array = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                if img_array is not None:
                    # Chuyen sang xam
                    gray = cv2.cvtColor(img_array, cv2.COLOR_BGR2GRAY)

                    # Tang kich thuoc anh de OCR tot hon
                    height, width = gray.shape
                    scale = max(2, int(min(800 / width, 800 / height)))
                    gray = cv2.resize(gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)

                    # Lam nhin hon bang Gaussian blur
                    gray = cv2.GaussianBlur(gray, (5, 5), 0)

                    # Adaptive threshold de tach text khoi nen
                    binary = cv2.adaptiveThreshold(
                        gray, 255,
                        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                        cv2.THRESH_BINARY,
                        11, 2
                    )

                    # OCR voi cau hinh tot hon
                    custom_config = r'--oem 3 --psm 6'
                    extracted_text = pytesseract.image_to_string(binary, lang='vie+eng', config=custom_config)
            except Exception as e:
                extracted_text = f'[OCR loi: {str(e)}]'
        else:
            extracted_text = '[OCR khong ho tro tren server nay]'

        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT COUNT(*) FROM invoices')
        count = c.fetchone()[0]
        conn.close()

        invoice_number = f"HD{datetime.date.today().strftime('%Y%m%d')}{count+1:04d}"

        return render_template('scan_result.html',
            invoice_number=invoice_number,
            image_data=image_data,
            extracted_text=extracted_text)

    return render_template('scan_image.html')

# Gui anh cho admin
@app.route('/invoice/<int:invoice_id>/send-to-admin', methods=['POST'])
@login_required
def send_to_admin(invoice_id):
    message = request.form.get('message', '')
    image_data = request.form.get('image_data', '')

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM users WHERE role = 'admin'")
    admin = c.fetchone()
    if admin:
        c.execute('''INSERT INTO notifications
            (user_id, message, invoice_id, is_read)
            VALUES (?,?,?,0)''',
            (admin['id'], message, invoice_id))
        conn.commit()
        flash('Da gui anh cho admin!', 'success')
    else:
        flash('Khong tim thay admin!', 'danger')
    conn.close()

    return redirect(url_for('invoices'))

@app.route('/notifications')
@login_required
def notifications():
    conn = get_db()
    c = conn.cursor()
    c.execute('''SELECT * FROM notifications
        WHERE user_id = ? ORDER BY created_at DESC LIMIT 50''',
        (session['user_id'],))
    notifs = c.fetchall()
    c.execute('UPDATE notifications SET is_read=1 WHERE user_id=?', (session['user_id'],))
    conn.commit()
    conn.close()
    return render_template('notifications.html', notifications=notifs)

@app.route('/invoice/<int:invoice_id>')
@login_required
def view_invoice(invoice_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM invoices WHERE id = ?', (invoice_id,))
    invoice = c.fetchone()
    conn.close()

    if not invoice:
        flash('Khong tim thay hoa don!', 'danger')
        return redirect(url_for('invoices'))

    # Parse items nhu list
    items_list = []
    if invoice.get('items'):
        try:
            items_list = json.loads(invoice['items'])
        except:
            items_list = []

    return render_template('invoice_view.html', invoice=invoice, items_list=items_list)

# Xuat Excel 1 hoa don chi tiet
@app.route('/invoice/<int:invoice_id>/export')
@login_required
def export_single_invoice(invoice_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM invoices WHERE id = ?', (invoice_id,))
    invoice = c.fetchone()
    conn.close()

    if not invoice:
        flash('Khong tim thay hoa don!', 'danger')
        return redirect(url_for('invoices'))

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoa Don"

    # Header
    ws['A1'] = 'HOA DON BAN HANG'
    ws['A1'].font = Font(bold=True, size=18, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 35

    # Info
    ws['A3'] = 'So hoa don:'
    ws['A3'].font = Font(bold=True)
    ws['B3'] = invoice['invoice_number']

    ws['A4'] = 'Ngay:'
    ws['A4'].font = Font(bold=True)
    ws['B4'] = invoice['date']

    ws['D3'] = 'Cua hang:'
    ws['D3'].font = Font(bold=True)
    ws.merge_cells('D3:D3')
    ws['E3'] = invoice['store_name'] or ''

    ws['D4'] = 'Nguoi tao:'
    ws['D4'].font = Font(bold=True)
    ws['E4'] = invoice['created_by']

    # Table header
    row = 6
    headers = ['STT', 'Ten Hang Hoa', 'So Luong', 'Don Gia', 'Thanh Tien']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    ws.row_dimensions[row].height = 25

    # Items
    import json as _json
    try:
        items = _json.loads(invoice['items']) if invoice['items'] else []
    except:
        items = []

    row = 7
    stt = 1
    for item in items:
        ws.cell(row=row, column=1, value=stt).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=item.get('name', ''))
        ws.cell(row=row, column=3, value=item.get('quantity', 1)).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4, value=item.get('price', 0))
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=5, value=item.get('total', 0))
        ws.cell(row=row, column=5).number_format = '#,##0'
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        row += 1
        stt += 1

    # Discount & Tax row
    if invoice['discount_percent'] > 0 or invoice['tax_percent'] > 0:
        if invoice['discount_percent'] > 0:
            disc_amt = invoice['subtotal'] * invoice['discount_percent'] / 100
            ws.cell(row=row, column=4, value=f'Giam gia ({invoice["discount_percent"]}%)')
            ws.cell(row=row, column=5, value=-disc_amt)
            ws.cell(row=row, column=5).number_format = '#,##0'
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            row += 1

        if invoice['tax_percent'] > 0:
            after_disc = invoice['subtotal'] * (1 - invoice['discount_percent']/100)
            tax_amt = after_disc * invoice['tax_percent'] / 100
            ws.cell(row=row, column=4, value=f'Thue ({invoice["tax_percent"]}%)')
            ws.cell(row=row, column=5, value=tax_amt)
            ws.cell(row=row, column=5).number_format = '#,##0'
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            row += 1

    # Total
    row += 1
    ws.cell(row=row, column=3, value='TONG CONG')
    ws.cell(row=row, column=3).font = Font(bold=True, size=12)
    ws.cell(row=row, column=5, value=invoice['total'])
    ws.cell(row=row, column=5).number_format = '#,##0'
    ws.cell(row=row, column=5).font = Font(bold=True, size=14, color='107C10')
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = PatternFill(
            start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        ws.cell(row=row, column=col).border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )

    # Ghi chu
    if invoice['notes']:
        row += 2
        ws.cell(row=row, column=1, value='Ghi chu:')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=invoice['notes'])
        ws.merge_cells(f'B{row}:E{row}')

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
        download_name=f'HoaDon_{invoice["invoice_number"]}.xlsx',
        as_attachment=True)

@app.route('/invoice/<int:invoice_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_invoice(invoice_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM invoices WHERE id = ?', (invoice_id,))
    invoice = c.fetchone()
    conn.close()

    if not invoice:
        flash('Khong tim thay hoa don!', 'danger')
        return redirect(url_for('invoices'))

    if request.method == 'POST':
        items_json = request.form.get('items_data', '[]')
        subtotal = float(request.form.get('subtotal', 0) or 0)
        discount = float(request.form.get('discount_percent', 0) or 0)
        tax = float(request.form.get('tax_percent', 0) or 0)
        total = subtotal * (1 - discount/100) * (1 + tax/100)

        image_data = invoice['image_data']
        if 'invoice_image' in request.files:
            file = request.files['invoice_image']
            if file and file.filename:
                img_bytes = file.read()
                image_data = base64.b64encode(img_bytes).decode('utf-8')

        conn = get_db()
        c = conn.cursor()
        c.execute('''UPDATE invoices SET
            invoice_number=?, date=?, store_name=?, items=?,
            subtotal=?, discount_percent=?, tax_percent=?, total=?,
            notes=?, image_data=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?''',
            (request.form.get('invoice_number'),
             request.form.get('invoice_date'),
             request.form.get('store_name'),
             items_json, subtotal, discount, tax, total,
             request.form.get('notes', ''),
             image_data,
             invoice_id))
        conn.commit()
        conn.close()

        flash('Cap nhat thanh cong!', 'success')
        return redirect(url_for('view_invoice', invoice_id=invoice_id))

    return render_template('invoice_form.html', invoice=invoice)

@app.route('/invoice/<int:invoice_id>/delete', methods=['POST'])
@login_required
def delete_invoice(invoice_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('DELETE FROM invoices WHERE id = ?', (invoice_id,))
    conn.commit()
    conn.close()
    flash('Xoa thanh cong!', 'success')
    return redirect(url_for('invoices'))

# ============================================================
# ROUTES - FINANCE
# ============================================================
@app.route('/finance')
@login_required
def finance():
    month = request.args.get('month', datetime.date.today().strftime('%Y-%m'))

    conn = get_db()
    c = conn.cursor()

    c.execute('''SELECT * FROM finances WHERE type='revenue' AND date LIKE ?
        ORDER BY date DESC''', (f'{month}%',))
    revenues = c.fetchall()

    c.execute('''SELECT * FROM finances WHERE type='expense' AND date LIKE ?
        ORDER BY date DESC''', (f'{month}%',))
    expenses = c.fetchall()

    total_rev = sum(r['amount'] for r in revenues)
    total_exp = sum(e['amount'] for e in expenses)

    conn.close()

    return render_template('finance.html',
        revenues=revenues, expenses=expenses,
        total_rev=total_rev, total_exp=total_exp,
        profit=total_rev - total_exp, month=month)

@app.route('/finance/add', methods=['POST'])
@login_required
def finance_add():
    conn = get_db()
    c = conn.cursor()
    c.execute('''INSERT INTO finances
        (type, date, amount, category, reason, description, created_by)
        VALUES (?,?,?,?,?,?,?)''',
        (request.form.get('type'),
         request.form.get('date'),
         float(request.form.get('amount', 0) or 0),
         request.form.get('category', ''),
         request.form.get('reason', ''),
         request.form.get('description', ''),
         session['username']))
    conn.commit()
    conn.close()
    flash('Them giao dich thanh cong!', 'success')
    return redirect(url_for('finance'))

@app.route('/finance/delete/<int:finance_id>', methods=['POST'])
@login_required
def finance_delete(finance_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('DELETE FROM finances WHERE id = ?', (finance_id,))
    conn.commit()
    conn.close()
    flash('Xoa thanh cong!', 'success')
    return redirect(url_for('finance'))

# ============================================================
# ROUTES - REPORT
# ============================================================
@app.route('/report')
@login_required
def report():
    date_from = request.args.get('date_from', datetime.date.today().replace(day=1).strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', datetime.date.today().strftime('%Y-%m-%d'))
    month = request.args.get('month', '')

    conn = get_db()
    c = conn.cursor()

    if month:
        date_from = f'{month}-01'
        date_to = f'{month}-31'

    c.execute('''SELECT * FROM invoices WHERE date >= ? AND date <= ?
        ORDER BY date DESC''', (date_from, date_to))
    invoices_list = c.fetchall()

    c.execute('''SELECT * FROM finances WHERE date >= ? AND date <= ?
        ORDER BY date DESC''', (date_from, date_to))
    finances = c.fetchall()

    total_inv = sum(i['total'] for i in invoices_list)
    total_rev = sum(f['amount'] for f in finances if f['type'] == 'revenue')
    total_exp = sum(f['amount'] for f in finances if f['type'] == 'expense')

    conn.close()

    return render_template('report.html',
        invoices=invoices_list, finances=finances,
        date_from=date_from, date_to=date_to, month=month,
        total_inv=total_inv, total_rev=total_rev,
        total_exp=total_exp, profit=total_rev - total_exp)

# ============================================================
# EXPORT INVOICES EXCEL
# ============================================================
@app.route('/invoices/export')
@login_required
def export_invoices():
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoa Don"

    title = 'DANH SACH HOA DON'
    if date_from and date_to:
        title = f'DANH SACH HOA DON - TU {date_from} DEN {date_to}'

    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 35

    headers = ['STT', 'So HD', 'Ngay', 'Cua Hang', 'Tong Tien', 'Ghi Chu', 'Nguoi Tao']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

    conn = get_db()
    c = conn.cursor()
    query = 'SELECT invoice_number, date, store_name, total, notes, created_by FROM invoices WHERE 1=1'
    params = []
    if date_from:
        query += ' AND date >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND date <= ?'
        params.append(date_to)
    query += ' ORDER BY date DESC'

    c.execute(query, params)
    row_num = 3
    stt = 1
    grand_total = 0

    for inv in c.fetchall():
        ws.cell(row=row_num, column=1, value=stt).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=inv[0])
        ws.cell(row=row_num, column=3, value=inv[1])
        ws.cell(row=row_num, column=4, value=inv[2] or '')
        ws.cell(row=row_num, column=5, value=inv[3])
        ws.cell(row=row_num, column=5).number_format = '#,##0'
        ws.cell(row=row_num, column=6, value=inv[4] or '')
        ws.cell(row=row_num, column=7, value=inv[5])
        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
        grand_total += float(inv[3] or 0)
        row_num += 1
        stt += 1

    # Total row
    ws.cell(row=row_num, column=1, value='TONG CONG').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=row_num, column=5, value=grand_total)
    ws.cell(row=row_num, column=5).number_format = '#,##0'
    ws.cell(row=row_num, column=5).font = Font(bold=True, color='FFFFFF')
    for col in range(1, 8):
        ws.cell(row=row_num, column=col).fill = PatternFill(
            start_color='0078D4', end_color='0078D4', fill_type='solid')
        ws.cell(row=row_num, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

    conn.close()

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'DanhSachHoaDon_{date_from or "all"}_{date_to or "all"}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True)

# ============================================================
# EXPORT FINANCE EXCEL
# ============================================================
@app.route('/finance/export')
@login_required
def export_finance():
    month = request.args.get('month', datetime.date.today().strftime('%Y-%m'))

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()

    conn = get_db()
    c = conn.cursor()

    # ===================== SHEET 1: DOANH THU =====================
    ws_rev = wb.active
    ws_rev.title = "1-Doanh Thu"
    ws_rev['A1'] = f'DOANH THU THANG {month}'
    ws_rev['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws_rev['A1'].fill = PatternFill(start_color='107C10', end_color='107C10', fill_type='solid')
    ws_rev['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_rev.merge_cells('A1:E1')
    ws_rev.row_dimensions[1].height = 35

    for i, h in enumerate(['STT', 'Ngay', 'Loai', 'So Tien', 'Mo Ta'], 1):
        cell = ws_rev.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='107C10', end_color='107C10', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

    c.execute('''SELECT date, category, amount, description FROM finances
        WHERE type='revenue' AND date LIKE ? ORDER BY date DESC''', (f'{month}%',))

    row_num = 3
    stt = 1
    total_rev = 0

    for rev in c.fetchall():
        ws_rev.cell(row=row_num, column=1, value=stt).alignment = Alignment(horizontal='center')
        ws_rev.cell(row=row_num, column=2, value=rev[0])
        ws_rev.cell(row=row_num, column=3, value=rev[1] or '')
        ws_rev.cell(row=row_num, column=4, value=rev[2])
        ws_rev.cell(row=row_num, column=4).number_format = '#,##0'
        ws_rev.cell(row=row_num, column=5, value=rev[3] or '')
        for col in range(1, 6):
            ws_rev.cell(row=row_num, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
        total_rev += float(rev[2] or 0)
        row_num += 1
        stt += 1

    ws_rev.cell(row=row_num, column=1, value='TONG DOANH THU').font = Font(bold=True, color='FFFFFF')
    ws_rev.cell(row=row_num, column=4, value=total_rev)
    ws_rev.cell(row=row_num, column=4).number_format = '#,##0'
    ws_rev.cell(row=row_num, column=4).font = Font(bold=True, color='FFFFFF')
    for col in range(1, 6):
        ws_rev.cell(row=row_num, column=col).fill = PatternFill(
            start_color='107C10', end_color='107C10', fill_type='solid')
        ws_rev.cell(row=row_num, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

    ws_rev.column_dimensions['A'].width = 8
    ws_rev.column_dimensions['B'].width = 14
    ws_rev.column_dimensions['C'].width = 20
    ws_rev.column_dimensions['D'].width = 18
    ws_rev.column_dimensions['E'].width = 30

    # ===================== SHEET 2: CHI PHI (CO LY DO) =====================
    ws_exp = wb.create_sheet("2-Chi Phi")
    ws_exp['A1'] = f'CHI PHI THANG {month}'
    ws_exp['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws_exp['A1'].fill = PatternFill(start_color='D13438', end_color='D13438', fill_type='solid')
    ws_exp['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_exp.merge_cells('A1:F1')
    ws_exp.row_dimensions[1].height = 35

    for i, h in enumerate(['STT', 'Ngay', 'Loai Chi Phi', 'So Tien', 'Ly Do Chi (Chi Nhung Gi?)', 'Mo Ta'], 1):
        cell = ws_exp.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='D13438', end_color='D13438', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

    c.execute('''SELECT date, category, amount, reason, description FROM finances
        WHERE type='expense' AND date LIKE ? ORDER BY date DESC''', (f'{month}%',))

    row_num = 3
    stt = 1
    total_exp = 0

    for exp in c.fetchall():
        ws_exp.cell(row=row_num, column=1, value=stt).alignment = Alignment(horizontal='center')
        ws_exp.cell(row=row_num, column=2, value=exp[0])
        ws_exp.cell(row=row_num, column=3, value=exp[1] or '')
        ws_exp.cell(row=row_num, column=4, value=exp[2])
        ws_exp.cell(row=row_num, column=4).number_format = '#,##0'
        ws_exp.cell(row=row_num, column=5, value=exp[3] or '')
        ws_exp.cell(row=row_num, column=6, value=exp[4] or '')
        for col in range(1, 7):
            ws_exp.cell(row=row_num, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
        total_exp += float(exp[2] or 0)
        row_num += 1
        stt += 1

    ws_exp.cell(row=row_num, column=1, value='TONG CHI PHI').font = Font(bold=True, color='FFFFFF')
    ws_exp.cell(row=row_num, column=4, value=total_exp)
    ws_exp.cell(row=row_num, column=4).number_format = '#,##0'
    ws_exp.cell(row=row_num, column=4).font = Font(bold=True, color='FFFFFF')
    for col in range(1, 7):
        ws_exp.cell(row=row_num, column=col).fill = PatternFill(
            start_color='D13438', end_color='D13438', fill_type='solid')
        ws_exp.cell(row=row_num, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

    ws_exp.column_dimensions['A'].width = 8
    ws_exp.column_dimensions['B'].width = 14
    ws_exp.column_dimensions['C'].width = 18
    ws_exp.column_dimensions['D'].width = 18
    ws_exp.column_dimensions['E'].width = 30
    ws_exp.column_dimensions['F'].width = 30

    # ===================== SHEET 3: TONG HOP =====================
    ws_sum = wb.create_sheet("3-Tong Hop")
    ws_sum['A1'] = f'TONG HOP TAI CHINH THANG {month}'
    ws_sum['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws_sum['A1'].fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
    ws_sum['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.merge_cells('A1:B1')
    ws_sum.row_dimensions[1].height = 35

    loi = total_rev - total_exp
    data = [
        ['CHI TIEU', 'SO TIEN'],
        ['Tong Doanh Thu (Thu)', total_rev],
        ['Tong Chi Phi (Chi)', total_exp],
        ['LOI NHUAN (Thu - Chi)', loi],
    ]

    for r, row_data in enumerate(data, 2):
        ws_sum.cell(row=r, column=1, value=row_data[0])
        ws_sum.cell(row=r, column=2, value=row_data[1])
        ws_sum.cell(row=r, column=2).number_format = '#,##0'

        if r == 2:
            ws_sum.cell(row=r, column=1).font = Font(bold=True, color='FFFFFF')
            ws_sum.cell(row=r, column=2).font = Font(bold=True, color='FFFFFF')
            ws_sum.cell(row=r, column=1).fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
            ws_sum.cell(row=r, column=2).fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
        elif r == 3:
            ws_sum.cell(row=r, column=1).font = Font(bold=True, color='107C10')
            ws_sum.cell(row=r, column=2).font = Font(bold=True, color='107C10')
        elif r == 4:
            ws_sum.cell(row=r, column=1).font = Font(bold=True, color='D13438')
            ws_sum.cell(row=r, column=2).font = Font(bold=True, color='D13438')
        elif r == 5:
            color = '107C10' if loi >= 0 else 'D13438'
            bg = 'E8F5E9' if loi >= 0 else 'FFEBEE'
            ws_sum.cell(row=r, column=1).font = Font(bold=True, size=14)
            ws_sum.cell(row=r, column=2).font = Font(bold=True, size=14, color=color)
            ws_sum.cell(row=r, column=1).fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            ws_sum.cell(row=r, column=2).fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')

        for col in range(1, 3):
            ws_sum.cell(row=r, column=col).border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='medium'))
            ws_sum.cell(row=r, column=col).alignment = Alignment(
                horizontal='left' if col == 1 else 'right',
                vertical='center')
            ws_sum.row_dimensions[r].height = 25

    ws_sum.column_dimensions['A'].width = 35
    ws_sum.column_dimensions['B'].width = 20

    conn.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f'TaiChinh_{month}.xlsx', as_attachment=True)

# ============================================================
# EXPORT REPORT EXCEL
# ============================================================
@app.route('/report/export')
@login_required
def report_export():
    date_from = request.args.get('date_from', datetime.date.today().replace(day=1).strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', datetime.date.today().strftime('%Y-%m-%d'))
    month = request.args.get('month', '')

    if month:
        date_from = f'{month}-01'
        date_to = f'{month}-31'

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    conn = get_db()
    c = conn.cursor()

    # SHEET 1: HOA DON
    ws = wb.active
    ws.title = "1-Hoa Don"
    title = f'BAO CAO HOA DON - TU {date_from} DEN {date_to}'
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 35

    headers = ['STT', 'So HD', 'Ngay', 'Cua Hang', 'Tong Tien', 'Ghi Chu', 'Nguoi Tao']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

    c.execute('''SELECT invoice_number, date, store_name, total, notes, created_by
        FROM invoices WHERE date >= ? AND date <= ? ORDER BY date DESC''',
        (date_from, date_to))

    row_num = 3
    stt = 1
    grand_total = 0

    for inv in c.fetchall():
        ws.cell(row=row_num, column=1, value=stt).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=inv[0])
        ws.cell(row=row_num, column=3, value=inv[1])
        ws.cell(row=row_num, column=4, value=inv[2] or '')
        ws.cell(row=row_num, column=5, value=inv[3])
        ws.cell(row=row_num, column=5).number_format = '#,##0'
        ws.cell(row=row_num, column=6, value=inv[4] or '')
        ws.cell(row=row_num, column=7, value=inv[5])
        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
        grand_total += float(inv[3] or 0)
        row_num += 1
        stt += 1

    ws.cell(row=row_num, column=1, value='TONG CONG HOA DON').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=row_num, column=5, value=grand_total)
    ws.cell(row=row_num, column=5).number_format = '#,##0'
    ws.cell(row=row_num, column=5).font = Font(bold=True, color='FFFFFF')
    for col in range(1, 8):
        ws.cell(row=row_num, column=col).fill = PatternFill(
            start_color='0078D4', end_color='0078D4', fill_type='solid')
        ws.cell(row=row_num, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15

    # SHEET 2: TAI CHINH
    ws2 = wb.create_sheet("2-Tai Chinh")

    ws2['A1'] = f'DOANH THU - TU {date_from} DEN {date_to}'
    ws2['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws2['A1'].fill = PatternFill(start_color='107C10', end_color='107C10', fill_type='solid')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.merge_cells('A1:E1')
    ws2.row_dimensions[1].height = 30

    for i, h in enumerate(['STT', 'Ngay', 'Loai', 'So Tien', 'Mo Ta'], 1):
        cell = ws2.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='107C10', end_color='107C10', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

    c.execute('''SELECT date, category, amount, description FROM finances
        WHERE type='revenue' AND date >= ? AND date <= ? ORDER BY date DESC''',
        (date_from, date_to))

    row_num = 3
    stt = 1
    total_rev = 0

    for rev in c.fetchall():
        ws2.cell(row=row_num, column=1, value=stt).alignment = Alignment(horizontal='center')
        ws2.cell(row=row_num, column=2, value=rev[0])
        ws2.cell(row=row_num, column=3, value=rev[1] or '')
        ws2.cell(row=row_num, column=4, value=rev[2])
        ws2.cell(row=row_num, column=4).number_format = '#,##0'
        ws2.cell(row=row_num, column=5, value=rev[3] or '')
        for col in range(1, 6):
            ws2.cell(row=row_num, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
        total_rev += float(rev[2] or 0)
        row_num += 1
        stt += 1

    row_num += 1
    ws2.cell(row=row_num, column=1, value='CHI PHI').font = Font(bold=True, size=13, color='FFFFFF')
    ws2.cell(row=row_num, column=1).fill = PatternFill(start_color='D13438', end_color='D13438', fill_type='solid')
    ws2.merge_cells(f'A{row_num}:E{row_num}')
    ws2.row_dimensions[row_num].height = 25

    row_num += 1
    for i, h in enumerate(['STT', 'Ngay', 'Loai Chi Phi', 'So Tien', 'Ly Do'], 1):
        cell = ws2.cell(row=row_num, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='D13438', end_color='D13438', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

    c.execute('''SELECT date, category, amount, reason FROM finances
        WHERE type='expense' AND date >= ? AND date <= ? ORDER BY date DESC''',
        (date_from, date_to))

    row_num += 1
    stt = 1
    total_exp = 0

    for exp in c.fetchall():
        ws2.cell(row=row_num, column=1, value=stt).alignment = Alignment(horizontal='center')
        ws2.cell(row=row_num, column=2, value=exp[0])
        ws2.cell(row=row_num, column=3, value=exp[1] or '')
        ws2.cell(row=row_num, column=4, value=exp[2])
        ws2.cell(row=row_num, column=4).number_format = '#,##0'
        ws2.cell(row=row_num, column=5, value=exp[3] or '')
        for col in range(1, 6):
            ws2.cell(row=row_num, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
        total_exp += float(exp[2] or 0)
        row_num += 1
        stt += 1

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 30

    # SHEET 3: TONG HOP
    ws3 = wb.create_sheet("3-Tong Hop")
    ws3['A1'] = f'TONG HOP BAO CAO - TU {date_from} DEN {date_to}'
    ws3['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws3['A1'].fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws3.merge_cells('A1:B1')
    ws3.row_dimensions[1].height = 35

    loi = total_rev - total_exp
    data = [
        ['CHI TIEU', 'SO TIEN'],
        ['Tong Hoa Don', grand_total],
        ['Tong Doanh Thu (Thu)', total_rev],
        ['Tong Chi Phi (Chi)', total_exp],
        ['LOI NHUAN', loi],
    ]

    for r, row_data in enumerate(data, 2):
        ws3.cell(row=r, column=1, value=row_data[0])
        ws3.cell(row=r, column=2, value=row_data[1])
        ws3.cell(row=r, column=2).number_format = '#,##0'

        if r == 2:
            ws3.cell(row=r, column=1).font = Font(bold=True, color='FFFFFF')
            ws3.cell(row=r, column=2).font = Font(bold=True, color='FFFFFF')
            ws3.cell(row=r, column=1).fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
            ws3.cell(row=r, column=2).fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
        elif r == 3:
            ws3.cell(row=r, column=1).font = Font(bold=True, color='0078D4')
            ws3.cell(row=r, column=2).font = Font(bold=True, color='0078D4')
        elif r == 4:
            ws3.cell(row=r, column=1).font = Font(bold=True, color='107C10')
            ws3.cell(row=r, column=2).font = Font(bold=True, color='107C10')
        elif r == 5:
            ws3.cell(row=r, column=1).font = Font(bold=True, color='D13438')
            ws3.cell(row=r, column=2).font = Font(bold=True, color='D13438')
        elif r == 6:
            color = '107C10' if loi >= 0 else 'D13438'
            bg = 'E8F5E9' if loi >= 0 else 'FFEBEE'
            ws3.cell(row=r, column=1).font = Font(bold=True, size=14)
            ws3.cell(row=r, column=2).font = Font(bold=True, size=14, color=color)
            ws3.cell(row=r, column=1).fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            ws3.cell(row=r, column=2).fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')

        for col in range(1, 3):
            ws3.cell(row=r, column=col).border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='medium'))
            ws3.cell(row=r, column=col).alignment = Alignment(
                horizontal='left' if col == 1 else 'right', vertical='center')
            ws3.row_dimensions[r].height = 25

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 20

    conn.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
        download_name=f'BaoCao_{date_from}_{date_to}.xlsx',
        as_attachment=True)

# ============================================================
# USERS
# ============================================================
@app.route('/users')
@admin_required
def users():
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, username, role, created_at FROM users ORDER BY created_at')
    users_list = c.fetchall()
    conn.close()
    return render_template('users.html', users=users_list)

@app.route('/users/add', methods=['POST'])
@admin_required
def users_add():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    role = request.form.get('role', 'ketoan')

    if not username or not password:
        flash('Vui long dien day du!', 'danger')
        return redirect(url_for('users'))

    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM users WHERE username = ?', (username,))
    if c.fetchone()[0] > 0:
        flash('Ten dang nhap da ton tai!', 'danger')
        conn.close()
        return redirect(url_for('users'))

    c.execute('INSERT INTO users (username, password, role) VALUES (?,?,?)',
              (username, generate_password_hash(password), role))
    conn.commit()
    conn.close()
    flash(f'Tao tai khoan {username} thanh cong!', 'success')
    return redirect(url_for('users'))

@app.route('/users/delete/<int:user_id>', methods=['POST'])
@admin_required
def users_delete(user_id):
    if user_id == session['user_id']:
        flash('Khong the xoa chinh minh!', 'danger')
        return redirect(url_for('users'))

    conn = get_db()
    c = conn.cursor()
    c.execute('DELETE FROM users WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()
    flash('Xoa thanh cong!', 'success')
    return redirect(url_for('users'))

# ============================================================
# BACKUP ALL DATA
# ============================================================
@app.route('/backup')
@login_required
def backup_data():
    import json
    from datetime import datetime as dt

    conn = get_db()
    c = conn.cursor()

    # Backup invoices
    c.execute('SELECT * FROM invoices ORDER BY date DESC')
    cols_inv = [d[0] for d in c._cursor.description]
    invoices = c.fetchall()
    invoices_list = [dict(zip(cols_inv, r)) for r in invoices]

    # Backup finances
    c.execute('SELECT * FROM finances ORDER BY date DESC')
    cols_fin = [d[0] for d in c._cursor.description]
    finances = c.fetchall()
    finances_list = [dict(zip(cols_fin, r)) for r in finances]

    # Backup users (khong lay password)
    c.execute('SELECT id, username, role, created_at FROM users')
    cols_user = [d[0] for d in c._cursor.description]
    users = c.fetchall()
    users_list = [dict(zip(cols_user, r)) for r in users]

    conn.close()

    # Tao JSON backup
    backup = {
        'backup_date': dt.now().isoformat(),
        'version': '1.0',
        'invoices': invoices_list,
        'finances': finances_list,
        'users': users_list
    }

    response = app.response_class(
        response=json.dumps(backup, indent=2, ensure_ascii=False),
        mimetype='application/json'
    )
    filename = f'backup_hoadon_{dt.now().strftime("%Y%m%d_%H%M%S")}.json'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response

# ============================================================
# ROUTES - NHAP HANG
# ============================================================

# Trang tong hop nhap hang - hien thi 2 muc: Tong hop NCC + Ton kho
@app.route('/import')
@login_required
def import_page():
    month = request.args.get('month', datetime.date.today().strftime('%Y-%m'))

    conn = get_db()
    c = conn.cursor()

    # Lay danh sach NCC
    c.execute('SELECT id, name FROM suppliers ORDER BY name')
    suppliers = c.fetchall()

    # Lay danh sach nguyen lieu (co nhom)
    c.execute('''SELECT m.id, m.name, m.unit, m.group_id, g.name as group_name, g.color as group_color
        FROM materials m LEFT JOIN material_groups g ON m.group_id = g.id
        ORDER BY g.sort_order, g.name, m.name''')
    materials = c.fetchall()

    # Lay danh sach nhom hang
    c.execute('SELECT * FROM material_groups ORDER BY sort_order, name')
    groups = c.fetchall()

    # Tong hop nhap theo NCC trong thang
    ncc_summary = []
    for sup in suppliers:
        c.execute('''SELECT COALESCE(SUM(i.total_price), 0) as total_debt,
            COALESCE(SUM(i.paid_amount), 0) as total_paid
            FROM imports i WHERE i.supplier_id = ? AND i.date LIKE ?''',
            (sup['id'], f'{month}%'))
        row = c.fetchone()
        total_debt = float(row[0]) if row and row[0] else 0
        total_paid = float(row[1]) if row and row[1] else 0
        if total_debt > 0:
            ncc_summary.append({
                'id': sup['id'], 'name': sup['name'],
                'total_debt': total_debt,
                'total_paid': total_paid,
                'remaining': total_debt - total_paid
            })

    # Tong hop theo nhom hang trong thang - luon hien thi tat ca nhom
    group_summary = []
    for grp in groups:
        c.execute('''SELECT COALESCE(SUM(i.total_price), 0) as total, COUNT(*) as cnt
            FROM imports i WHERE i.material_id IN
            (SELECT id FROM materials WHERE group_id = ?) AND i.date LIKE ?''',
            (grp['id'], f'{month}%'))
        row = c.fetchone()
        total = float(row[0]) if row and row[0] else 0
        cnt = int(row[1]) if row and row[1] else 0
        group_summary.append({
            'id': grp['id'], 'name': grp['name'],
            'color': grp['color'] or '#6c757d', 'total_amount': total, 'count': cnt
        })
    # Nhom "Khong co nhom"
    c.execute('''SELECT COALESCE(SUM(i.total_price), 0) as total, COUNT(*) as cnt
        FROM imports i WHERE i.material_id IN
        (SELECT id FROM materials WHERE group_id IS NULL) AND i.date LIKE ?''',
        (f'{month}%',))
    row = c.fetchone()
    group_summary.append({
        'id': 0, 'name': 'Khong co nhom',
        'color': '#6c757d',
        'total_amount': float(row[0]) if row and row[0] else 0,
        'count': int(row[1]) if row and row[1] else 0
    })

    # Lay tat ca import trong thang
    c.execute('''SELECT i.id, i.date, s.name as supplier_name, s.id as supplier_id,
        m.name as material_name, m.id as material_id,
        COALESCE(NULLIF(TRIM(i.unit), ''), m.unit) as unit,
        i.quantity, i.unit_price, i.total_price, i.paid_amount,
        COALESCE(i.payment_count, 0) as payment_count,
        i.paid_details,
        COALESCE(i.paid_amount, 0) as paid_amount_val
        FROM imports i
        LEFT JOIN suppliers s ON i.supplier_id = s.id
        LEFT JOIN materials m ON i.material_id = m.id
        WHERE i.date LIKE ? ORDER BY i.date DESC, i.id DESC''',
        (f'{month}%',))
    imports_list = c.fetchall()

    # Tinh ton kho cho tung nguyen lieu trong thang
    # opening = closing_stock cua thang truoc
    # import trong thang
    # closing = opening + import - xuat
    # xuat = (opening + import) - closing  ->  xuat = import + opening - closing

    prev_month = (datetime.date(int(month[:4]), int(month[5:7]), 1)
                  - datetime.timedelta(days=1)).strftime('%Y-%m')
    next_year = int(month[:4])
    next_mon = int(month[5:7]) + 1
    if next_mon > 12:
        next_year += 1
        next_mon = 1
    next_month = f'{next_year:04d}-{next_mon:02d}'

    inventory_data = []
    for mat in materials:
        # ton cuoi thang truoc
        c.execute('''SELECT closing_stock FROM inventory
            WHERE month = ? AND material_id = ?''',
            (prev_month, mat['id']))
        row = c.fetchone()
        opening = float(row[0]) if row and row[0] else 0

        # tong nhap trong thang
        c.execute('''SELECT COALESCE(SUM(quantity), 0) FROM imports
            WHERE date LIKE ? AND material_id = ?''',
            (f'{month}%', mat['id']))
        row = c.fetchone()
        import_qty = float(row[0]) if row and row[0] else 0

        # ton cuoi thang hien tai (co the chua cap nhat)
        c.execute('''SELECT closing_stock FROM inventory
            WHERE month = ? AND material_id = ?''',
            (month, mat['id']))
        row = c.fetchone()
        closing = float(row[0]) if row and row[0] else 0

        # xuat = (nhap + ton thang truoc) - ton thang hien tai
        export_qty = import_qty + opening - closing if closing > 0 else 0

        inventory_data.append({
            'material': mat,
            'opening': opening,
            'import_qty': import_qty,
            'export_qty': export_qty,
            'closing': closing,
        })

    conn.close()

    return render_template('import.html',
        suppliers=suppliers, materials=materials,
        ncc_summary=ncc_summary, imports_list=imports_list,
        inventory_data=inventory_data, month=month,
        today_str=datetime.date.today().strftime('%Y-%m-%d'),
        prev_month_url=prev_month, next_month_url=next_month,
        group_summary=group_summary)


# Trang nhap nguyen lieu nhanh - chi de nhap, khong xem bang
@app.route('/import/enter', methods=['GET', 'POST'])
@login_required
def import_enter():
    conn = get_db()
    c = conn.cursor()

    # Lay danh sach nguyen lieu (co nhom)
    c.execute('''SELECT m.id, m.name, m.unit, m.group_id,
        g.name as group_name, g.color as group_color
        FROM materials m LEFT JOIN material_groups g ON m.group_id = g.id
        ORDER BY g.sort_order, g.name, m.name''')
    materials = c.fetchall()

    # Lay danh sach NCC
    c.execute('SELECT id, name FROM suppliers ORDER BY name')
    suppliers = c.fetchall()

    # Lay nhom hang de loc
    c.execute('SELECT * FROM material_groups ORDER BY sort_order, name')
    groups = c.fetchall()

    today_str = datetime.date.today().strftime('%Y-%m-%d')
    default_supplier_id = request.args.get('supplier_id', '')
    default_material_id = request.args.get('material_id', '')

    # Lay phieu nhap gan day
    c.execute('''SELECT i.date, s.name as supplier_name, m.name as material_name,
        i.quantity, i.unit_price, i.total_price, i.id
        FROM imports i
        LEFT JOIN suppliers s ON i.supplier_id = s.id
        LEFT JOIN materials m ON i.material_id = m.id
        ORDER BY i.id DESC LIMIT 50''')
    recent_imports = c.fetchall()

    conn.close()

    if request.method == 'POST':
        date = request.form.get('date', today_str)
        material_id = request.form.get('material_id', '')
        supplier_id = request.form.get('supplier_id', '')
        quantity = float(request.form.get('quantity', 0) or 0)
        unit_price = float(request.form.get('unit_price', 0) or 0)
        notes = request.form.get('notes', '')

        if not material_id or quantity <= 0 or unit_price <= 0:
            flash('Vui long dien day du thong tin!', 'danger')
            conn.close()
            return render_template('import_enter.html',
                materials=materials, suppliers=suppliers, groups=groups,
                today_str=today_str,
                default_supplier_id=default_supplier_id,
                default_material_id=default_material_id)

        total_price = quantity * unit_price
        month = date[:7]

        c.execute('''INSERT INTO imports
            (date, material_id, supplier_id, quantity, unit_price, total_price, notes, created_by)
            VALUES (?,?,?,?,?,?,?,?)''',
            (date, material_id, supplier_id, quantity, unit_price, total_price, notes,
             session.get('username', 'admin')))

        # Cap nhat ton kho
        c.execute('''SELECT id, opening_stock, import_qty, closing_stock FROM inventory
            WHERE month = ? AND material_id = ?''',
            (month, material_id))
        inv = c.fetchone()
        if inv:
            new_import_qty = float(inv['import_qty'] or 0) + quantity
            new_closing = float(inv['opening_stock'] or 0) + new_import_qty
            c.execute('''UPDATE inventory SET import_qty=?, closing_stock=?,
                updated_at=CURRENT_TIMESTAMP WHERE id=?''',
                (new_import_qty, new_closing, inv['id']))
        else:
            prev_month = (datetime.date(int(month[:4]), int(month[5:7]), 1)
                          - datetime.timedelta(days=1)).strftime('%Y-%m')
            c.execute('''SELECT closing_stock FROM inventory
                WHERE month = ? AND material_id = ?''',
                (prev_month, material_id))
            prev = c.fetchone()
            opening = float(prev[0]) if prev and prev[0] else 0
            closing = opening + quantity
            c.execute('''INSERT INTO inventory
                (month, material_id, opening_stock, import_qty, closing_stock)
                VALUES (?,?,?,?,?)''',
                (month, material_id, opening, quantity, closing))

        conn.commit()
        flash(f'Da nhap thanh cong! {quantity} x {unit_price:,.0f} = {total_price:,.0f} VND', 'success')

        # Giu lai gia tri da nhap de nhap tiep
        default_supplier_id = supplier_id
        default_material_id = material_id

    conn.close()
    return render_template('import_enter.html',
        materials=materials, suppliers=suppliers, groups=groups,
        today_str=today_str,
        default_supplier_id=default_supplier_id,
        default_material_id=default_material_id,
        recent_imports=recent_imports)


# Them phieu nhap
@app.route('/import/add', methods=['POST'])
@login_required
def import_add():
    date = request.form.get('date', '')
    material_id = request.form.get('material_id', '')
    supplier_id = request.form.get('supplier_id', '')
    quantity = float(request.form.get('quantity', 0) or 0)
    unit_price = float(request.form.get('unit_price', 0) or 0)
    unit = request.form.get('unit', '').strip()
    paid_amount = float(request.form.get('paid_amount', 0) or 0)
    notes = request.form.get('notes', '')

    if not date or not material_id or quantity <= 0 or unit_price <= 0:
        flash('Vui long dien day du thong tin!', 'danger')
        return redirect(url_for('import_page'))

    total_price = quantity * unit_price
    month = date[:7]  # YYYY-MM

    # Tinh payment count va details
    payment_count = 0
    paid_details = ''
    if paid_amount > 0:
        payment_count = 1
        today_str = datetime.date.today().strftime('%Y-%m-%d')
        paid_details = f'{today_str}:{int(paid_amount)}'

    conn = get_db()
    c = conn.cursor()

    # Cap nhat don vi neu co
    if unit:
        c.execute('UPDATE materials SET unit=? WHERE id=?', (unit, material_id))

    # Insert phieu nhap
    if USE_PG:
        c.execute('''INSERT INTO imports
            (date, material_id, supplier_id, quantity, unit, unit_price, total_price,
             notes, paid_amount, payment_count, paid_details, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)''',
            (date, material_id, supplier_id, quantity, unit, unit_price, total_price,
             notes, paid_amount, payment_count, paid_details, session['username']))
    else:
        c.execute('''INSERT INTO imports
            (date, material_id, supplier_id, quantity, unit, unit_price, total_price,
             notes, paid_amount, payment_count, paid_details, created_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (date, material_id, supplier_id, quantity, unit, unit_price, total_price,
             notes, paid_amount, payment_count, paid_details, session['username']))

    # Cap nhat ton kho: neu chua co record thi insert, neu co thi update
    # Cong nhap vao closing_stock
    c.execute('''SELECT id, opening_stock, import_qty, closing_stock FROM inventory
        WHERE month = ? AND material_id = ?''',
        (month, material_id))
    inv = c.fetchone()

    if inv:
        new_import_qty = float(inv['import_qty'] or 0) + quantity
        new_closing = float(inv['opening_stock'] or 0) + new_import_qty
        c.execute('''UPDATE inventory SET import_qty=?, closing_stock=?,
            updated_at=CURRENT_TIMESTAMP WHERE id=?''',
            (new_import_qty, new_closing, inv['id']))
    else:
        # Tinh opening tu thang truoc
        prev_month = (datetime.date(int(month[:4]), int(month[5:7]), 1)
                      - datetime.timedelta(days=1)).strftime('%Y-%m')
        c.execute('''SELECT closing_stock FROM inventory
            WHERE month = ? AND material_id = ?''',
            (prev_month, material_id))
        prev = c.fetchone()
        opening = float(prev[0]) if prev and prev[0] else 0
        closing = opening + quantity

        c.execute('''INSERT INTO inventory
            (month, material_id, opening_stock, import_qty, closing_stock)
            VALUES (?,?,?,?,?)''',
            (month, material_id, opening, quantity, closing))

    conn.commit()
    conn.close()

    flash('Them phieu nhap thanh cong!', 'success')
    return redirect(url_for('import_page', month=month))


# Cap nhat ton kho (closing stock)
@app.route('/import/update-inventory', methods=['POST'])
@login_required
def import_update_inventory():
    month = request.form.get('month', '')
    material_id = request.form.get('material_id', '')
    closing = float(request.form.get('closing_stock', 0) or 0)

    conn = get_db()
    c = conn.cursor()

    c.execute('''SELECT id, opening_stock, import_qty FROM inventory
        WHERE month = ? AND material_id = ?''',
        (month, material_id))
    inv = c.fetchone()

    if inv:
        c.execute('''UPDATE inventory SET closing_stock=?, updated_at=CURRENT_TIMESTAMP WHERE id=?''',
            (closing, inv['id']))
    else:
        prev_month = (datetime.date(int(month[:4]), int(month[5:7]), 1)
                      - datetime.timedelta(days=1)).strftime('%Y-%m')
        c.execute('''SELECT closing_stock FROM inventory
            WHERE month = ? AND material_id = ?''',
            (prev_month, material_id))
        prev = c.fetchone()
        opening = float(prev[0]) if prev and prev[0] else 0
        c.execute('''INSERT INTO inventory (month, material_id, opening_stock, import_qty, closing_stock)
            VALUES (?,?,?,0,?)''',
            (month, material_id, opening, closing))

    conn.commit()
    conn.close()

    flash('Cap nhat ton kho thanh cong!', 'success')
    return redirect(url_for('import_page', month=month))


# Xoa phieu nhap
@app.route('/import/delete/<int:import_id>', methods=['POST'])
@login_required
def import_delete(import_id):
    conn = get_db()
    c = conn.cursor()

    # Lay thong tin phieu nhap
    c.execute('SELECT date, material_id, quantity FROM imports WHERE id = ?', (import_id,))
    imp = c.fetchone()
    if imp:
        month = imp['date'][:7]
        material_id = imp['material_id']

        # Tru quantity khoi ton kho
        c.execute('''SELECT id, opening_stock, import_qty, closing_stock FROM inventory
            WHERE month = ? AND material_id = ?''',
            (month, material_id))
        inv = c.fetchone()
        if inv:
            new_import_qty = max(0, float(inv['import_qty'] or 0) - float(imp['quantity'] or 0))
            new_closing = float(inv['opening_stock'] or 0) + new_import_qty
            c.execute('''UPDATE inventory SET import_qty=?, closing_stock=?,
                updated_at=CURRENT_TIMESTAMP WHERE id=?''',
                (new_import_qty, new_closing, inv['id']))

        c.execute('DELETE FROM imports WHERE id = ?', (import_id,))
        conn.commit()
        flash('Xoa phieu nhap thanh cong!', 'success')
    conn.close()
    return redirect(url_for('import_page', month=month or datetime.date.today().strftime('%Y-%m')))


# Xoa nhieu phieu nhap cung luc (xoa hang loat)
@app.route('/import/delete-batch', methods=['POST'])
@login_required
def import_delete_batch():
    ids_str = request.form.get('delete_ids_batch', '')
    month = request.form.get('month', datetime.date.today().strftime('%Y-%m'))

    if not ids_str:
        flash('Chua chon phieu nao de xoa!', 'warning')
        return redirect(url_for('import_page', month=month))

    try:
        ids = [int(x.strip()) for x in ids_str.split(',') if x.strip()]
    except ValueError:
        flash('Du lieu khong hop le!', 'danger')
        return redirect(url_for('import_page', month=month))

    if not ids:
        flash('Chua chon phieu nao de xoa!', 'warning')
        return redirect(url_for('import_page', month=month))

    conn = get_db()
    c = conn.cursor()

    for import_id in ids:
        c.execute('SELECT date, material_id, quantity FROM imports WHERE id = ?', (int(import_id),))
        imp = c.fetchone()
        if imp:
            imp_month = imp['date'][:7]
            material_id = imp['material_id']

            c.execute('''SELECT id, opening_stock, import_qty, closing_stock FROM inventory
                WHERE month = ? AND material_id = ?''',
                (imp_month, material_id))
            inv = c.fetchone()
            if inv:
                new_import_qty = max(0, float(inv['import_qty'] or 0) - float(imp['quantity'] or 0))
                new_closing = float(inv['opening_stock'] or 0) + new_import_qty
                c.execute('''UPDATE inventory SET import_qty=?, closing_stock=?,
                    updated_at=CURRENT_TIMESTAMP WHERE id=?''',
                    (new_import_qty, new_closing, inv['id']))

            c.execute('DELETE FROM imports WHERE id = ?', (int(import_id),))

    conn.commit()
    conn.close()
    flash(f'Da xoa {len(ids)} phieu nhap!', 'success')
    return redirect(url_for('import_page', month=month))


# Cap nhat so tien da thanh toan cho 1 phieu nhap
@app.route('/import/payment/<int:import_id>', methods=['POST'])
@login_required
def import_update_payment(import_id):
    paid = float(request.form.get('paid_amount', 0) or 0)
    month = request.form.get('month', datetime.date.today().strftime('%Y-%m'))
    conn = get_db()
    c = conn.cursor()
    c.execute('UPDATE imports SET paid_amount = ? WHERE id = ?', (paid, import_id))
    conn.commit()
    conn.close()
    flash('Cap nhat thanh toan thanh cong!', 'success')
    return redirect(url_for('import_page', month=month))


# Sua phieu nhap (cap nhat nhieu truong)
@app.route('/import/update-single', methods=['POST'])
@login_required
def import_update_single():
    import_id = int(request.form.get('import_id', 0))
    new_date = request.form.get('date', '').strip()
    new_supplier_id = request.form.get('supplier_id', '') or None
    new_material_id = request.form.get('material_id', '') or None
    new_quantity = float(request.form.get('quantity', 0) or 0)
    new_unit_price = float(request.form.get('unit_price', 0) or 0)
    new_unit = request.form.get('unit', '').strip()
    new_notes = request.form.get('notes', '').strip()
    new_paid = float(request.form.get('paid_amount', 0) or 0)
    month = request.form.get('month', datetime.date.today().strftime('%Y-%m'))

    if not new_date or not new_material_id:
        flash('Ngay va Nguyen lieu la bat buoc!', 'danger')
        return redirect(url_for('import_page', month=month))

    conn = get_db()
    c = conn.cursor()

    # Lay thong tin cu
    c.execute('SELECT date, material_id, quantity FROM imports WHERE id = ?', (import_id,))
    old = c.fetchone()
    if not old:
        conn.close()
        flash('Khong tim thay phieu nhap!', 'danger')
        return redirect(url_for('import_page', month=month))

    old_month = old['date'][:7]
    old_material_id = old['material_id']
    old_qty = float(old['quantity'] or 0)

    # Cap nhat ton kho thang cu (tru qty cu)
    if old_material_id:
        c.execute('''SELECT id, opening_stock, import_qty, closing_stock FROM inventory
            WHERE month = ? AND material_id = ?''',
            (old_month, old_material_id))
        inv = c.fetchone()
        if inv:
            new_import_qty = max(0, float(inv['import_qty'] or 0) - old_qty)
            new_closing = float(inv['opening_stock'] or 0) + new_import_qty
            c.execute('''UPDATE inventory SET import_qty=?, closing_stock=?,
                updated_at=CURRENT_TIMESTAMP WHERE id=?''',
                (new_import_qty, new_closing, inv['id']))

    new_total = new_quantity * new_unit_price
    new_month = new_date[:7]

    # Cap nhat ton kho thang moi (cong qty moi)
    if new_material_id:
        c.execute('''SELECT id, opening_stock, import_qty, closing_stock FROM inventory
            WHERE month = ? AND material_id = ?''',
            (new_month, new_material_id))
        inv2 = c.fetchone()
        if inv2:
            new_imp = float(inv2['import_qty'] or 0) + new_quantity
            new_close = float(inv2['opening_stock'] or 0) + new_imp
            c.execute('''UPDATE inventory SET import_qty=?, closing_stock=?,
                updated_at=CURRENT_TIMESTAMP WHERE id=?''',
                (new_imp, new_close, inv2['id']))
        else:
            # Tao ban ghi ton kho moi
            prev_month = (datetime.date(int(new_month[:4]), int(new_month[5:7]), 1)
                          - datetime.timedelta(days=1)).strftime('%Y-%m')
            c.execute('SELECT closing_stock FROM inventory WHERE month = ? AND material_id = ?',
                      (prev_month, new_material_id))
            prev = c.fetchone()
            opening = float(prev[0]) if prev and prev[0] else 0
            closing = opening + new_quantity
            c.execute('''INSERT INTO inventory (month, material_id, opening_stock, import_qty, closing_stock)
                VALUES (?,?,?,?,?)''',
                (new_month, new_material_id, opening, new_quantity, closing))

    # Cap nhat phieu nhap
    c.execute('''UPDATE imports SET date=?, material_id=?, supplier_id=?,
        quantity=?, unit=?, unit_price=?, total_price=?, notes=?, paid_amount=?
        WHERE id=?''',
        (new_date, new_material_id, new_supplier_id, new_quantity,
         new_unit, new_unit_price, new_total, new_notes, new_paid, import_id))

    conn.commit()
    conn.close()
    flash('Cap nhat phieu nhap thanh cong!', 'success')
    return redirect(url_for('import_page', month=month))


# Nhap hang tu text (batch import - giong import_text nhung cho nhap)
@app.route('/import/batch', methods=['GET', 'POST'])
@login_required
def import_batch():
    conn = get_db()
    c = conn.cursor()

    if request.method == 'POST':
        date = request.form.get('import_date', datetime.date.today().strftime('%Y-%m-%d'))
        supplier_id = request.form.get('supplier_id', '') or None
        group_id = request.form.get('group_id', '') or None
        month = date[:7]

        added = 0

        # Lay tat ca cac field tu form
        for key in sorted(request.form.keys()):
            if key.startswith('name_'):
                idx = key[5:]
                name = request.form.get('name_' + idx, '').strip()
                if not name:
                    continue
                qty = float(request.form.get('qty_' + idx, 0) or 0)
                unit = request.form.get('unit_' + idx, '').strip()
                price = float(request.form.get('price_' + idx, 0) or 0)
                total = qty * price

                # Lay nhom cua dong nay, hoac mac dinh tu global
                row_group_id = request.form.get('row_group_' + idx, '') or None
                effective_group_id = row_group_id if row_group_id else group_id

                # Tim hoac tao nguyen lieu
                c.execute('SELECT id, group_id FROM materials WHERE name = ?', (name,))
                mat = c.fetchone()
                if mat:
                    material_id = mat['id']
                    # Neu material chua co nhom, gan nhom cua dong nay
                    if (mat['group_id'] is None or mat['group_id'] == '') and effective_group_id:
                        c.execute('UPDATE materials SET group_id=? WHERE id=?', (effective_group_id, material_id))
                    # Cap nhat don vi (luc nao cung update, vi user co the nhap dung don vi)
                    if unit and unit.strip():
                        c.execute("UPDATE materials SET unit=? WHERE id=?",
                                  (unit, material_id))
                else:
                    if USE_PG:
                        c.execute('INSERT INTO materials (name, group_id, unit) VALUES (%s,%s,%s) RETURNING id',
                                  (name, effective_group_id, unit))
                        material_id = c.fetchone()[0]
                    else:
                        c.execute('INSERT INTO materials (name, group_id, unit) VALUES (?,?,?)',
                                  (name, effective_group_id, unit))
                        material_id = c.lastrowid

                if USE_PG:
                    c.execute('''INSERT INTO imports
                        (date, material_id, supplier_id, quantity, unit, unit_price, total_price, created_by)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)''',
                        (date, material_id, supplier_id, qty, unit, price, total, session['username']))
                else:
                    c.execute('''INSERT INTO imports
                        (date, material_id, supplier_id, quantity, unit, unit_price, total_price, created_by)
                        VALUES (?,?,?,?,?,?,?,?)''',
                        (date, material_id, supplier_id, qty, unit, price, total, session['username']))

                # Cap nhat ton kho
                c.execute('''SELECT id, opening_stock, import_qty, closing_stock FROM inventory
                    WHERE month = ? AND material_id = ?''',
                    (month, material_id))
                inv = c.fetchone()
                if inv:
                    new_import_qty = float(inv['import_qty'] or 0) + qty
                    new_closing = float(inv['opening_stock'] or 0) + new_import_qty
                    c.execute('''UPDATE inventory SET import_qty=?, closing_stock=?,
                        updated_at=CURRENT_TIMESTAMP WHERE id=?''',
                        (new_import_qty, new_closing, inv['id']))
                else:
                    prev_month = (datetime.date(int(month[:4]), int(month[5:7]), 1)
                                  - datetime.timedelta(days=1)).strftime('%Y-%m')
                    c.execute('''SELECT closing_stock FROM inventory
                        WHERE month = ? AND material_id = ?''',
                        (prev_month, material_id))
                    prev = c.fetchone()
                    opening = float(prev[0]) if prev and prev[0] else 0
                    closing = opening + qty
                    c.execute('''INSERT INTO inventory
                        (month, material_id, opening_stock, import_qty, closing_stock)
                        VALUES (?,?,?,?,?)''',
                        (month, material_id, opening, qty, closing))

                added += 1

        conn.commit()
        conn.close()
        flash(f'Da them {added} phieu nhap!', 'success')
        return redirect(url_for('import_batch', month=month))

    c.execute('SELECT id, name FROM suppliers ORDER BY name')
    suppliers = c.fetchall()
    c.execute('SELECT id, name FROM materials ORDER BY name')
    materials = c.fetchall()
    c.execute('SELECT * FROM material_groups ORDER BY sort_order, name')
    groups = c.fetchall()

    # Lay group summary hien tai (bao gom ca khong co nhom)
    today = datetime.date.today()
    current_month = today.strftime('%Y-%m')
    c.execute('''SELECT g.id, g.name, g.color,
        COALESCE(SUM(i.total_price), 0) as total_amount,
        COUNT(i.id) as count
        FROM material_groups g
        LEFT JOIN materials m ON m.group_id = g.id
        LEFT JOIN imports i ON i.material_id = m.id AND i.date LIKE ?
        GROUP BY g.id ORDER BY g.sort_order, g.name''', (f'{current_month}%',))
    group_summary = list(c.fetchall())

    # Lay du lieu khong co nhom
    c.execute('''SELECT
        COALESCE(SUM(i.total_price), 0) as total_amount,
        COUNT(i.id) as count
        FROM imports i
        LEFT JOIN materials m ON i.material_id = m.id
        WHERE i.date LIKE ? AND m.group_id IS NULL''',
        (f'{current_month}%',))
    no_group = c.fetchone()
    no_group_summary = {
        'id': 0,
        'name': 'Khong co nhom',
        'color': '#cccccc',
        'total_amount': no_group['total_amount'] if no_group else 0,
        'count': no_group['count'] if no_group else 0
    }
    conn.close()

    return render_template('import_batch.html',
        suppliers=suppliers, materials=materials, groups=groups,
        today_str=datetime.date.today().strftime('%Y-%m-%d'),
        group_summary=group_summary,
        no_group_summary=no_group_summary)


# Sua nhieu phieu cung luc - chi NCC va Nhom Hang
@app.route('/import/batch-edit', methods=['POST'])
@login_required
def import_batch_edit():
    ids_str = request.form.get('edit_ids', '')
    if not ids_str:
        flash('Chua chon phieu nao!', 'danger')
        return redirect(url_for('import_page'))

    ids = [int(x.strip()) for x in ids_str.split(',') if x.strip()]
    if not ids:
        flash('Danh sach ID khong hop le!', 'danger')
        return redirect(url_for('import_page'))

    month = request.form.get('month', datetime.date.today().strftime('%Y-%m'))
    new_supplier_id = request.form.get('supplier_id', '').strip()
    new_group_id = request.form.get('group_id', '').strip()

    conn = get_db()
    c = conn.cursor()

    updated = 0

    # Cap nhat NCC cho cac phieu
    if new_supplier_id:
        for import_id in ids:
            c.execute('UPDATE imports SET supplier_id=? WHERE id=?', (new_supplier_id, import_id))
            updated += 1

    # Chi update nhom cho material khi material CHUA CO nhom
    if new_group_id:
        for import_id in ids:
            c.execute('SELECT material_id FROM imports WHERE id = ?', (import_id,))
            row = c.fetchone()
            if row and row['material_id']:
                # Chi gan nhom khi material chua co nhom
                c.execute('SELECT group_id FROM materials WHERE id = ?', (row['material_id'],))
                mat_row = c.fetchone()
                mat_group = mat_row['group_id'] if mat_row else None
                if mat_group is None or mat_group == '':
                    c.execute('UPDATE materials SET group_id=? WHERE id=?', (new_group_id, row['material_id']))
                    updated += 1

    if not new_supplier_id and not new_group_id:
        flash('Chua co gi de cap nhat!', 'warning')
        conn.close()
        return redirect(url_for('import_page', month=month))

    conn.commit()
    conn.close()
    flash(f'Da cap nhat {updated} phieu!', 'success')
    return redirect(url_for('import_page', month=month))


# Xu ly thanh toan tu trang batch
@app.route('/import/pay', methods=['POST'])
@login_required
def import_pay():
    pay_date = request.form.get('pay_date', '').strip()
    supplier_id = request.form.get('supplier_id', '').strip()
    pay_amount = float(request.form.get('pay_amount', 0) or 0)
    pay_note = request.form.get('pay_note', '').strip()
    month = request.form.get('month', datetime.date.today().strftime('%Y-%m'))

    if not pay_date or not supplier_id or pay_amount <= 0:
        flash('Vui long dien day du thong tin thanh toan!', 'danger')
        return redirect(url_for('import_batch', month=month))

    conn = get_db()
    c = conn.cursor()

    # Lay cac phieu chua tra het cua NCC nay
    c.execute('''SELECT id, total_price, paid_amount
        FROM imports
        WHERE supplier_id = ? AND (paid_amount IS NULL OR paid_amount < total_price)
        ORDER BY date ASC''',
        (int(supplier_id),))
    unpaid = list(c.fetchall())

    if not unpaid:
        flash('Khong co phieu nao chua thanh toan cho NCC nay!', 'warning')
        conn.close()
        return redirect(url_for('import_batch', month=month))

    remaining = pay_amount
    updated_count = 0
    payment_details = []

    for row in unpaid:
        import_id = row['id']
        total_price = float(row['total_price'] or 0)
        current_paid = float(row['paid_amount'] or 0)
        need = total_price - current_paid

        if need <= 0:
            continue

        if remaining >= need:
            # Tra du
            remaining -= need
            new_paid = total_price
        else:
            # Tra thieu
            new_paid = current_paid + remaining
            remaining = 0

        # Cap nhat paid_details
        c.execute('SELECT paid_details, payment_count FROM imports WHERE id = ?', (import_id,))
        old = c.fetchone()
        old_details = old['paid_details'] or ''
        old_count = old['payment_count'] or 0

        # Tach details cu, them payment moi
        details_parts = []
        if old_details:
            details_parts = old_details.split(';')
        details_parts.append(f'{pay_date}:{int(new_paid - current_paid)}')

        new_count = old_count + 1

        c.execute('''UPDATE imports SET paid_amount = ?, payment_count = ?,
            paid_details = ? WHERE id = ?''',
            (new_paid, new_count, ';'.join(details_parts), import_id))
        updated_count += 1
        payment_details.append(f'{new_paid - current_paid}')

        if remaining <= 0:
            break

    conn.commit()
    conn.close()

    if updated_count > 0:
        paid_str = pay_amount - remaining
        flash(f'Da thanh toan {updated_count} phieu, tong {int(paid_str):,} VND!', 'success')
    else:
        flash('Khong co phieu nao duoc cap nhat!', 'warning')

    return redirect(url_for('import_batch', month=month))


# Export Excel phan Nhap
@app.route('/import/export')
@login_required
def export_import():
    month = request.args.get('month', datetime.date.today().strftime('%Y-%m'))

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()

    conn = get_db()
    c = conn.cursor()

    # Lay du lieu import (co nhom hang)
    c.execute('''SELECT i.date, s.name as supplier_name, m.name as material_name,
        g.name as group_name, g.color as group_color,
        m.unit, i.quantity, i.unit_price, i.total_price,
        COALESCE(i.paid_amount, 0) as paid_amount
        FROM imports i
        LEFT JOIN suppliers s ON i.supplier_id = s.id
        LEFT JOIN materials m ON i.material_id = m.id
        LEFT JOIN material_groups g ON m.group_id = g.id
        WHERE i.date LIKE ? ORDER BY i.date ASC, i.id ASC''',
        (f'{month}%',))
    imports_list = c.fetchall()

    # Lay du lieu NCC
    c.execute('''SELECT s.name,
        COALESCE(SUM(i.total_price), 0) as total_debt,
        COALESCE(SUM(i.paid_amount), 0) as total_paid
        FROM imports i
        LEFT JOIN suppliers s ON i.supplier_id = s.id
        WHERE i.date LIKE ? AND s.id IS NOT NULL
        GROUP BY s.id, s.name
        ORDER BY s.name''',
        (f'{month}%',))
    ncc_rows = c.fetchall()

    # Lay du lieu ton kho
    c.execute('''SELECT m.name, m.unit, i.opening_stock, i.import_qty,
        i.closing_stock
        FROM inventory i
        LEFT JOIN materials m ON i.material_id = m.id
        WHERE i.month = ? ORDER BY m.name''',
        (month,))
    inv_rows = c.fetchall()

    # Lay du lieu tong hop theo nhom hang
    c.execute('''SELECT g.name as group_name, g.color as group_color,
        COALESCE(SUM(i.total_price), 0) as total, COUNT(i.id) as cnt
        FROM imports i
        LEFT JOIN materials m ON i.material_id = m.id
        LEFT JOIN material_groups g ON m.group_id = g.id
        WHERE i.date LIKE ?
        GROUP BY g.id, g.name, g.color
        ORDER BY g.sort_order, g.name''',
        (f'{month}%',))
    group_rows = c.fetchall()

    # ==================== SHEET 1: CHI TIET NHAP HANG ====================
    ws = wb.active
    ws.title = "1-Chi Tiet Nhap"

    ws['A1'] = f'CHI TIET NHAP HANG THANG {month}'
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 35

    headers = ['Ngay', 'Nha Cung Cap', 'Ten Nguyen Lieu', 'Nhom Hang', 'Don Vi', 'So Luong', 'Don Gia', 'Thanh Tien', 'Da Thanh Toan', 'Con Lai']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = 3
    grand_total = 0
    total_paid = 0
    for imp in imports_list:
        paid = float(imp[9]) if len(imp) > 9 and imp[9] else 0
        remaining = float(imp[8] or 0) - paid
        grand_total += float(imp[8] or 0)
        total_paid += paid
        ws.cell(row=row_num, column=1, value=imp[0])
        ws.cell(row=row_num, column=2, value=imp[1] or '')
        ws.cell(row=row_num, column=3, value=imp[2] or '')
        ws.cell(row=row_num, column=4, value=imp[3] or '')
        ws.cell(row=row_num, column=5, value=imp[5] or '')
        ws.cell(row=row_num, column=6, value=float(imp[6] or 0))
        ws.cell(row=row_num, column=6).number_format = '#,##0.00'
        ws.cell(row=row_num, column=7, value=float(imp[7] or 0))
        ws.cell(row=row_num, column=7).number_format = '#,##0'
        ws.cell(row=row_num, column=8, value=float(imp[8] or 0))
        ws.cell(row=row_num, column=8).number_format = '#,##0'
        ws.cell(row=row_num, column=9, value=paid)
        ws.cell(row=row_num, column=9).number_format = '#,##0'
        ws.cell(row=row_num, column=10, value=remaining)
        ws.cell(row=row_num, column=10).number_format = '#,##0'
        for col in range(1, 11):
            ws.cell(row=row_num, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
        row_num += 1

    # Tong cong
    ws.cell(row=row_num, column=1, value='TONG CONG').font = Font(bold=True, color='FFFFFF')
    for col in range(1, 11):
        ws.cell(row=row_num, column=col).fill = PatternFill(
            start_color='0078D4', end_color='0078D4', fill_type='solid')
        ws.cell(row=row_num, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(row=row_num, column=8, value=grand_total)
    ws.cell(row=row_num, column=8).number_format = '#,##0'
    ws.cell(row=row_num, column=8).font = Font(bold=True, color='FFFFFF')
    ws.cell(row=row_num, column=9, value=total_paid)
    ws.cell(row=row_num, column=9).number_format = '#,##0'
    ws.cell(row=row_num, column=9).font = Font(bold=True, color='FFFFFF')
    ws.cell(row=row_num, column=10, value=grand_total - total_paid)
    ws.cell(row=row_num, column=10).number_format = '#,##0'
    ws.cell(row=row_num, column=10).font = Font(bold=True, color='FFFFFF')

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18

    # ==================== SHEET 2: TONG HOP NCC ====================
    ws2 = wb.create_sheet("2-Tong Hop NCC")

    ws2['A1'] = f'TONG HOP CONG NO NCC THANG {month}'
    ws2['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws2['A1'].fill = PatternFill(start_color='107C10', end_color='107C10', fill_type='solid')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.merge_cells('A1:E1')
    ws2.row_dimensions[1].height = 35

    for i, h in enumerate(['Nha Cung Cap', 'Tong No', 'Da Thanh Toan', 'Con Lai', 'Ghi Chu'], 1):
        cell = ws2.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='107C10', end_color='107C10', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = 3
    total_no = 0
    total_paid = 0
    total_remaining = 0
    for ncc in ncc_rows:
        debt = float(ncc[1])
        paid = float(ncc[2]) if ncc[2] else 0
        remaining = debt - paid
        total_no += debt
        total_paid += paid
        total_remaining += remaining
        ws2.cell(row=row_num, column=1, value=ncc[0])
        ws2.cell(row=row_num, column=2, value=debt)
        ws2.cell(row=row_num, column=2).number_format = '#,##0'
        ws2.cell(row=row_num, column=3, value=paid)
        ws2.cell(row=row_num, column=3).number_format = '#,##0'
        ws2.cell(row=row_num, column=4, value=remaining)
        ws2.cell(row=row_num, column=4).number_format = '#,##0'
        ws2.cell(row=row_num, column=5, value=f'Tong nhap thang {month}')
        for col in range(1, 6):
            ws2.cell(row=row_num, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
        row_num += 1

    # Tong cong
    ws2.cell(row=row_num, column=1, value='TONG CONG').font = Font(bold=True, color='FFFFFF')
    ws2.cell(row=row_num, column=1).fill = PatternFill(start_color='107C10', end_color='107C10', fill_type='solid')
    ws2.cell(row=row_num, column=2, value=total_no)
    ws2.cell(row=row_num, column=2).number_format = '#,##0'
    ws2.cell(row=row_num, column=2).font = Font(bold=True, color='FFFFFF')
    ws2.cell(row=row_num, column=3, value=total_paid)
    ws2.cell(row=row_num, column=3).number_format = '#,##0'
    ws2.cell(row=row_num, column=3).font = Font(bold=True, color='FFFFFF')
    ws2.cell(row=row_num, column=4, value=total_remaining)
    ws2.cell(row=row_num, column=4).number_format = '#,##0'
    ws2.cell(row=row_num, column=4).font = Font(bold=True, color='FFFFFF')
    for col in range(1, 6):
        ws2.cell(row=row_num, column=col).fill = PatternFill(start_color='107C10', end_color='107C10', fill_type='solid')
        ws2.cell(row=row_num, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 25

    # ==================== SHEET 3: TON KHO ====================
    ws3 = wb.create_sheet("3-Ton Kho")

    ws3['A1'] = f'TON KHO THANG {month}'
    ws3['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws3['A1'].fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws3.merge_cells('A1:F1')
    ws3.row_dimensions[1].height = 35

    for i, h in enumerate(['Ten Nguyen Lieu', 'Don Vi', 'Ton Dau Thang', 'Nhap', 'Xuat', 'Ton Cuoi Thang'], 1):
        cell = ws3.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
        cell.fill = PatternFill(start_color='D48806', end_color='D48806', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = 3
    for inv in inv_rows:
        ws3.cell(row=row_num, column=1, value=inv[0] or '')
        ws3.cell(row=row_num, column=2, value=inv[1] or '')
        ws3.cell(row=row_num, column=3, value=float(inv[2] or 0))
        ws3.cell(row=row_num, column=3).number_format = '#,##0.00'
        ws3.cell(row=row_num, column=4, value=float(inv[3] or 0))
        ws3.cell(row=row_num, column=4).number_format = '#,##0.00'
        export_val = float(inv[2] or 0) + float(inv[3] or 0) - float(inv[4] or 0)
        ws3.cell(row=row_num, column=5, value=export_val if float(inv[4] or 0) > 0 else 0)
        ws3.cell(row=row_num, column=5).number_format = '#,##0.00'
        ws3.cell(row=row_num, column=6, value=float(inv[4] or 0))
        ws3.cell(row=row_num, column=6).number_format = '#,##0.00'
        for col in range(1, 7):
            ws3.cell(row=row_num, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
        row_num += 1

    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 12
    ws3.column_dimensions['F'].width = 15

    # ==================== SHEET 4: TONG HOP NHOM HANG ====================
    ws4 = wb.create_sheet("4-Tong Hop Nhom")

    ws4['A1'] = f'TONG HOP NHOM HANG THANG {month}'
    ws4['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws4['A1'].fill = PatternFill(start_color='6c3F00', end_color='6c3F00', fill_type='solid')
    ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws4.merge_cells('A1:E1')
    ws4.row_dimensions[1].height = 35

    for i, h in enumerate(['Nhom Hang', 'So Phieu', 'Tong Gia Tri', 'Ty Le (%)', 'Ghi Chu'], 1):
        cell = ws4.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='6c3F00', end_color='6c3F00', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = 3
    grand_total = sum(float(r[2] or 0) for r in group_rows)
    for grp in group_rows:
        total = float(grp[2] or 0)
        pct = (total / grand_total * 100) if grand_total > 0 else 0
        note = f'Tong nhap thang {month}'
        ws4.cell(row=row_num, column=1, value=grp[0] or 'Khong co nhom')
        ws4.cell(row=row_num, column=2, value=int(grp[3] or 0))
        ws4.cell(row=row_num, column=3, value=total)
        ws4.cell(row=row_num, column=3).number_format = '#,##0'
        ws4.cell(row=row_num, column=4, value=pct / 100)
        ws4.cell(row=row_num, column=4).number_format = '0.0%'
        ws4.cell(row=row_num, column=5, value=note)
        for col in range(1, 6):
            ws4.cell(row=row_num, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
        row_num += 1

    # Tong cong
    ws4.cell(row=row_num, column=1, value='TONG CONG').font = Font(bold=True, color='FFFFFF')
    ws4.cell(row=row_num, column=2, value=sum(int(r[3] or 0) for r in group_rows))
    ws4.cell(row=row_num, column=3, value=grand_total)
    ws4.cell(row=row_num, column=3).number_format = '#,##0'
    ws4.cell(row=row_num, column=4, value=1.0)
    ws4.cell(row=row_num, column=4).number_format = '0.0%'
    for col in range(1, 6):
        ws4.cell(row=row_num, column=col).fill = PatternFill(
            start_color='6c3F00', end_color='6c3F00', fill_type='solid')
        ws4.cell(row=row_num, column=col).font = Font(bold=True, color='FFFFFF')
        ws4.cell(row=row_num, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 12
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 12
    ws4.column_dimensions['E'].width = 25

    conn.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f'NhapHang_{month}.xlsx', as_attachment=True)


# ==================== QUAN LY NCC & NGUYEN LIEU ====================

@app.route('/import/suppliers')
@login_required
def suppliers_page():
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM suppliers ORDER BY name')
    suppliers_list = c.fetchall()
    conn.close()
    return render_template('suppliers.html', suppliers=suppliers_list)


@app.route('/import/suppliers/add', methods=['POST'])
@login_required
def suppliers_add():
    name = request.form.get('name', '').strip()
    phone = request.form.get('phone', '').strip()
    address = request.form.get('address', '').strip()
    notes = request.form.get('notes', '').strip()

    if not name:
        flash('Vui long nhap ten NCC!', 'danger')
        return redirect(url_for('suppliers_page'))

    conn = get_db()
    c = conn.cursor()
    c.execute('INSERT INTO suppliers (name, phone, address, notes) VALUES (?,?,?,?)',
              (name, phone, address, notes))
    conn.commit()
    conn.close()
    flash(f'Them NCC "{name}" thanh cong!', 'success')
    return redirect(url_for('suppliers_page'))


@app.route('/import/suppliers/delete/<int:supplier_id>', methods=['POST'])
@login_required
def suppliers_delete(supplier_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('DELETE FROM suppliers WHERE id = ?', (supplier_id,))
    conn.commit()
    conn.close()
    flash('Xoa NCC thanh cong!', 'success')
    return redirect(url_for('suppliers_page'))


@app.route('/import/materials')
@login_required
def materials_page():
    conn = get_db()
    c = conn.cursor()
    c.execute('''SELECT m.*, g.name as group_name, g.color as group_color
        FROM materials m LEFT JOIN material_groups g ON m.group_id = g.id
        ORDER BY g.sort_order, g.name, m.name''')
    materials_list = c.fetchall()
    c.execute('SELECT * FROM material_groups ORDER BY sort_order, name')
    groups = c.fetchall()
    conn.close()
    return render_template('materials.html', materials=materials_list, groups=groups)


@app.route('/import/materials/add', methods=['POST'])
@login_required
def materials_add():
    name = request.form.get('name', '').strip()
    unit = request.form.get('unit', '').strip()
    group_id = request.form.get('group_id', '')

    if not name:
        flash('Vui long nhap ten nguyen lieu!', 'danger')
        return redirect(url_for('materials_page'))

    conn = get_db()
    c = conn.cursor()
    c.execute('INSERT INTO materials (name, unit, group_id) VALUES (?,?,?)',
              (name, unit, group_id or None))
    conn.commit()
    conn.close()
    flash(f'Them nguyen lieu "{name}" thanh cong!', 'success')
    return redirect(url_for('materials_page'))


@app.route('/import/materials/delete/<int:material_id>', methods=['POST'])
@login_required
def materials_delete(material_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('DELETE FROM materials WHERE id = ?', (material_id,))
    conn.commit()
    conn.close()
    flash('Xoa nguyen lieu thanh cong!', 'success')
    return redirect(url_for('materials_page'))


# ==================== QUAN LY NHOM HANG ====================

@app.route('/import/groups')
@login_required
def groups_page():
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM material_groups ORDER BY sort_order, name')
    groups = c.fetchall()
    # Dem so nguyen lieu trong tung nhom
    group_counts = {}
    c.execute('SELECT group_id, COUNT(*) as cnt FROM materials GROUP BY group_id')
    for row in c.fetchall():
        group_counts[row['group_id'] or 0] = row['cnt']
    conn.close()
    return render_template('groups.html', groups=groups, group_counts=group_counts)


@app.route('/import/groups/add', methods=['POST'])
@login_required
def groups_add():
    name = request.form.get('name', '').strip()
    color = request.form.get('color', '#0078D4').strip()
    sort_order = int(request.form.get('sort_order', 0) or 0)

    if not name:
        flash('Vui long nhap ten nhom!', 'danger')
        return redirect(url_for('groups_page'))

    conn = get_db()
    c = conn.cursor()
    c.execute('INSERT INTO material_groups (name, color, sort_order) VALUES (?,?,?)',
              (name, color, sort_order))
    conn.commit()
    conn.close()
    flash(f'Them nhom "{name}" thanh cong!', 'success')
    return redirect(url_for('groups_page'))


@app.route('/import/groups/delete/<int:group_id>', methods=['POST'])
@login_required
def groups_delete(group_id):
    conn = get_db()
    c = conn.cursor()
    # Chuyen nguyen lieu ve khong nhom
    c.execute('UPDATE materials SET group_id = NULL WHERE group_id = ?', (group_id,))
    c.execute('DELETE FROM material_groups WHERE id = ?', (group_id,))
    conn.commit()
    conn.close()
    flash('Xoa nhom thanh cong!', 'success')
    return redirect(url_for('groups_page'))


# ==================== CHI TIET NCC ====================

@app.route('/import/supplier/<int:supplier_id>')
@login_required
def supplier_detail(supplier_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM suppliers WHERE id = ?', (supplier_id,))
    supplier = c.fetchone()
    if not supplier:
        flash('Khong tim thay NCC!', 'danger')
        return redirect(url_for('import_page'))

    # Lay tat ca phieu nhap cua NCC
    c.execute('''SELECT i.date, m.name as material_name, m.unit, i.quantity,
        i.unit_price, i.total_price, i.notes, i.id, COALESCE(i.paid_amount, 0) as paid_amount
        FROM imports i
        LEFT JOIN materials m ON i.material_id = m.id
        WHERE i.supplier_id = ?
        ORDER BY i.date DESC, i.id DESC''',
        (supplier_id,))
    imports = c.fetchall()

    total_debt = sum(float(i['total_price'] or 0) for i in imports)
    total_paid = sum(float(i['paid_amount'] or 0) for i in imports)

    # Theo thang
    monthly = {}
    for imp in imports:
        m = imp['date'][:7]
        if m not in monthly:
            monthly[m] = {'debt': 0, 'paid': 0}
        monthly[m]['debt'] += float(imp['total_price'] or 0)
        monthly[m]['paid'] += float(imp['paid_amount'] or 0)

    conn.close()
    return render_template('supplier_detail.html',
        supplier=supplier, imports=imports,
        total_debt=total_debt, total_paid=total_paid, monthly=monthly)


# ==================== CHI TIET NHOM HANG ====================

@app.route('/import/group/<int:group_id>')
@login_required
def group_detail(group_id):
    if group_id == 0:
        # Hien thi tat ca nguyen lieu khong co nhom
        group_name = 'Khong co nhom'
        group_color = '#6c757d'
    else:
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT * FROM material_groups WHERE id = ?', (group_id,))
        group = c.fetchone()
        if not group:
            flash('Khong tim thay nhom!', 'danger')
            return redirect(url_for('import_page'))
        group_name = group['name']
        group_color = group['color']
        conn.close()

    conn = get_db()
    c = conn.cursor()
    if group_id == 0:
        c.execute('''SELECT m.*, g.name as group_name, g.color as group_color
            FROM materials m LEFT JOIN material_groups g ON m.group_id = g.id
            WHERE m.group_id IS NULL ORDER BY m.name''')
    else:
        c.execute('''SELECT m.*, g.name as group_name, g.color as group_color
            FROM materials m LEFT JOIN material_groups g ON m.group_id = g.id
            WHERE m.group_id = ? ORDER BY m.name''',
            (group_id,))
    materials_list = c.fetchall()

    # Lay phieu nhap cho tung nguyen lieu trong nhom
    material_ids = [str(m['id']) for m in materials_list]
    material_imports = {}
    for mat in materials_list:
        material_imports[mat['id']] = []

    if material_ids:
        placeholders = ','.join(['?' for _ in material_ids])
        c.execute(f'''SELECT i.date, m.name as material_name, s.name as supplier_name,
            i.quantity, i.unit_price, i.total_price, i.material_id, i.id
            FROM imports i
            LEFT JOIN materials m ON i.material_id = m.id
            LEFT JOIN suppliers s ON i.supplier_id = s.id
            WHERE i.material_id IN ({placeholders})
            ORDER BY i.date DESC, i.id DESC''',
            material_ids)
        for row in c.fetchall():
            mid = row['material_id']
            if mid in material_imports:
                material_imports[mid].append(row)

    total_all = 0
    mat_totals = {}  # luu tong tien moi nguyen lieu
    for mat in materials_list:
        mat_total = sum(float(r['total_price'] or 0) for r in material_imports.get(mat['id'], []))
        mat_totals[mat['id']] = mat_total
        total_all += mat_total

    conn.close()
    return render_template('group_detail.html',
        group_id=group_id, group_name=group_name, group_color=group_color or '#6c757d',
        materials_list=materials_list, material_imports=material_imports,
        total_all=total_all, mat_totals=mat_totals)


# ==================== CHI TIET NGUYEN LIEU ====================

@app.route('/import/material/<int:material_id>')
@login_required
def material_detail(material_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('''SELECT m.*, g.name as group_name, g.color as group_color
        FROM materials m LEFT JOIN material_groups g ON m.group_id = g.id
        WHERE m.id = ?''', (material_id,))
    material = c.fetchone()
    if not material:
        flash('Khong tim thay nguyen lieu!', 'danger')
        return redirect(url_for('materials_page'))

    # Lay phieu nhap
    c.execute('''SELECT i.date, s.name as supplier_name, i.quantity,
        i.unit_price, i.total_price, i.notes, i.id
        FROM imports i
        LEFT JOIN suppliers s ON i.supplier_id = s.id
        WHERE i.material_id = ?
        ORDER BY i.date DESC, i.id DESC''',
        (material_id,))
    imports = c.fetchall()

    total_import = sum(float(i['total_price'] or 0) for i in imports)

    # Theo thang
    monthly = {}
    for imp in imports:
        m = imp['date'][:7]
        if m not in monthly:
            monthly[m] = {'qty': 0, 'amount': 0}
        monthly[m]['qty'] += float(imp['quantity'] or 0)
        monthly[m]['amount'] += float(imp['total_price'] or 0)

    conn.close()
    return render_template('material_detail.html',
        material=material, imports=imports,
        total_import=total_import, monthly=monthly)


# ============================================================
# ERROR HANDLERS
# ============================================================
@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('500.html'), 500

# ============================================================
# MAIN
# ============================================================
init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
